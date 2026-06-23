"""
runner.py
=========
Runs Hybrid GA-ALNS for the first 100 products found in data/inventory_flow.csv.
Outputs are written to output/ folder:
  - schedule.csv          : per-product per-warehouse per-period schedule
  - costs.csv             : per-product per-warehouse per-period cost breakdown
  - cost_summary.csv      : one row per product with total + component costs
  - run_log.csv           : timing and fitness per product

Usage:
    python runner.py
    python runner.py --max-products 10
    python runner.py --config model/config.json
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import math
import random
import sys
import time
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level  = logging.INFO,
    format = "%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt= "%H:%M:%S",
)
log = logging.getLogger("runner")

# ---------------------------------------------------------------------------
# Internal imports
# ---------------------------------------------------------------------------
from model.core.problem      import Problem
from model.core.objective    import ObjectiveCalculator
from model.core.constraints  import ConstraintHandler
from model.core.decoder      import Decoder
from model.ga.operators      import GeneticOperators
from model.ga.genetic_algorithm import GeneticAlgorithm
from model.ga.milp_seed      import generate_milp_seed
from model.alns.alns_solver  import ALNSSolver

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
def load_config(config_path: Path) -> Dict[str, Any]:
    return json.loads(config_path.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def load_inventory_flow(n: int = 10000) -> Tuple[List[str], Dict]:
    """Return first n product_ids from inventory_flow.csv plus the flow data."""
    path = DATA_DIR / "inventory_flow.csv"
    product_order: List[str] = []
    flow: Dict[str, Dict[str, Dict[int, Dict]]] = defaultdict(lambda: defaultdict(dict))

    with path.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = row["product_id"]
            wh  = row["warehouse_id"]
            t   = int(row["time_period"])
            if pid not in product_order:
                if len(product_order) >= n:
                    continue
                product_order.append(pid)
            if pid in product_order:
                flow[pid][wh][t] = {
                    "delta_I": float(row["inventory_fluctuation"]),
                    "U"      : float(row["inventory_ceiling"]),
                    "L"      : float(row["inventory_floor"]),
                }
    return product_order, flow


def load_inventory_begin() -> Dict[str, Dict[str, float]]:
    bi: Dict[str, Dict[str, float]] = defaultdict(dict)
    with (DATA_DIR / "inventory_begin.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            bi[row["product_id"]][row["warehouse_id"]] = float(row["beginning_inventory"])
    return bi


def load_unit_costs() -> Dict[str, Dict[str, Dict[int, Dict]]]:
    costs: Dict = defaultdict(lambda: defaultdict(dict))
    with (DATA_DIR / "unit_cost.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            pid = row["product_id"]
            wh  = row["warehouse_id"]
            t   = int(row["time_period"])
            costs[pid][wh][t] = {
                "Co": float(row["overstock_cost"]),
                "Cs": float(row["shortage_cost"]),
                "Cb": float(row["backlog_cost"]),
                "Cp": float(row["penalty_cost"]),
            }
    return costs


def load_vendor_capacity() -> Dict[str, Dict[int, float]]:
    cap: Dict[str, Dict[int, float]] = defaultdict(dict)
    with (DATA_DIR / "vendor_capacity.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cap[row["product_id"]][int(row["time_period"])] = float(row["capacity"])
    return cap


def load_packing_details() -> Dict[str, int]:
    cp: Dict[str, int] = {}
    with (DATA_DIR / "packing_details.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cp[row["product_id"]] = int(row["pack_multiple"])
    return cp


def load_oa_lead_times() -> Dict[str, int]:
    path = DATA_DIR / "oa_lead_time.csv"
    lt: Dict[str, int] = {}
    with path.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            wh = _format_wh(row["warehouse_id"])
            lt[wh] = int(row["week_lead_time"])
    return lt


def _format_wh(val: str) -> str:
    val = val.strip()
    if val.startswith("WH"):
        return val
    try:
        return f"WH{int(val):02d}"
    except ValueError:
        return val


def load_plt_lead_times() -> Dict[Tuple[str, str], int]:
    lt: Dict[Tuple[str, str], int] = {}
    with (DATA_DIR / "plt_lead_time.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            frm = _format_wh(row["from_warehouse_id"])
            to  = _format_wh(row["to_warehouse_id"])
            lt[(frm, to)] = int(row["lead_time_weeks"])
    return lt


def load_distances() -> Dict[Tuple[str, str], float]:
    """
    FGPs_distance.csv has State as row/col label.
    WH01=MI, WH02=OH, WH03=IN, WH04=IL, WH05=KY, WH06=MO
    """
    state_map = {
        "MI": "WH01", "OH": "WH02", "IN": "WH03",
        "IL": "WH04", "KY": "WH05", "MO": "WH06",
    }
    dist: Dict[Tuple[str, str], float] = {}
    with (DATA_DIR / "FGPs_distance.csv").open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            from_state = row["State"].strip()
            from_wh = state_map.get(from_state)
            if not from_wh:
                continue
            for state, wh in state_map.items():
                if state in row:
                    dist[(from_wh, wh)] = float(row[state])
    return dist


# ---------------------------------------------------------------------------
# Problem factory from CSV data
# ---------------------------------------------------------------------------

def build_problem(
    product_id  : str,
    flow        : Dict[str, Dict[int, Dict]],   # wh -> t -> {delta_I, U, L}
    bi_map      : Dict[str, float],             # wh -> BI
    costs_map   : Dict[str, Dict[int, Dict]],   # wh -> t -> {Co,Cs,Cb,Cp}
    cap_map     : Dict[int, float],             # t -> CAP
    cp          : int,
    lt_oa       : int | Dict[str, int],
    lt_plt_map  : Dict[Tuple[str, str], int],
    dist_map    : Dict[Tuple[str, str], float],
    tc          : float = 1.2,
) -> Problem:
    warehouses = tuple(sorted(flow.keys()))
    periods_set = set()
    for wh_data in flow.values():
        periods_set.update(wh_data.keys())
    periods = tuple(sorted(periods_set))

    # Filter PLT lead times to only warehouses in this problem
    LT_PLT = {(i, j): v for (i, j), v in lt_plt_map.items() if i in warehouses and j in warehouses}

    if isinstance(lt_oa, dict):
        LT_OA = {wh: int(lt_oa[wh]) for wh in warehouses}
    else:
        LT_OA = {wh: int(lt_oa) for wh in warehouses}

    PLT_periods = {
        wh: frozenset(t for t in periods if t <= LT_OA[wh] - 1)
        for wh in warehouses
    }

    BI      = {wh: bi_map.get(wh, 0.0) for wh in warehouses}
    delta_I = {}
    U = {}
    L = {}
    for wh in warehouses:
        for t in periods:
            entry = flow[wh].get(t, {})
            delta_I[(wh, t)] = entry.get("delta_I", 0.0)
            U[(wh, t)] = entry.get("U", 1e6)
            L[(wh, t)] = entry.get("L", 0.0)

    CAP = {t: cap_map.get(t, 9999.0) for t in periods}

    Co = {}
    Cs = {}
    Cb = {}
    Cp_map = {}
    for wh in warehouses:
        for t in periods:
            c = costs_map.get(wh, {}).get(t, {})
            Co[(wh, t)] = c.get("Co", 0.1)
            Cs[(wh, t)] = c.get("Cs", 0.5)
            Cb[(wh, t)] = c.get("Cb", 1500.0)
            Cp_map[(wh, t)] = c.get("Cp", 2000.0)

    dist = {(i, j): dist_map.get((i, j), 0.0) for i in warehouses for j in warehouses}

    Cp_plt = {}
    for (i, j) in LT_PLT:
        for t in periods:
            Cp_plt[(i, j, t)] = Cp_map.get((j, t), 0.0)

    return Problem(
        product     = product_id,
        warehouses  = warehouses,
        periods     = periods,
        LT_OA       = LT_OA,
        LT_PLT      = LT_PLT,
        PLT_periods = PLT_periods,
        BI          = BI,
        delta_I     = delta_I,
        U           = U,
        L           = L,
        CAP         = CAP,
        CP          = cp,
        TC          = tc,
        dist        = dist,
        Co          = Co,
        Cs          = Cs,
        Cb          = Cb,
        Cp          = Cp_map,
        Cp_plt      = Cp_plt,
    )


# ---------------------------------------------------------------------------
# Component factory
# ---------------------------------------------------------------------------

def build_components(problem: Problem, cfg: Dict, rng: random.Random):
    ga_cfg   = cfg["ga"]
    alns_cfg = cfg["alns"]
    stop_cfg = cfg["stopping"]

    obj_calc   = ObjectiveCalculator(problem)
    constraint = ConstraintHandler(problem)
    decoder    = Decoder(problem, obj_calc)

    alns = ALNSSolver(
        problem        = problem,
        decoder        = decoder,
        obj_calc       = obj_calc,
        constraint     = constraint,
        n_iterations   = alns_cfg["n_iterations"],
        q_min_ratio    = alns_cfg["q_min_ratio"],
        q_max_ratio    = alns_cfg["q_max_ratio"],
        lambda_rho     = alns_cfg["lambda_rho"],
        segment_size   = alns_cfg["segment_size"],
        sa_accept_prob = alns_cfg["sa_initial_accept_prob"],
        sa_cooling     = alns_cfg["sa_cooling"],
        rng            = rng,
    )

    operators = GeneticOperators(
        problem     = problem,
        constraint  = constraint,
        p_crossover = ga_cfg["p_crossover"],
        p_mutation  = ga_cfg["p_mutation"],
        rng         = rng,
    )

    milp_seed = None
    if ga_cfg.get("milp_seed_fraction", 0) > 0:
        milp_seed = generate_milp_seed(
            problem    = problem,
            time_limit = ga_cfg.get("milp_time_limit", 15),
        )

    def _log_cb(gen: int, best_f: float) -> None:
        if gen % 50 == 0:
            log.info("    Gen %4d | best = %.4f", gen, best_f)

    ga = GeneticAlgorithm(
        problem      = problem,
        decoder      = decoder,
        obj_calc     = obj_calc,
        constraint   = constraint,
        operators    = operators,
        alns_solver  = alns,
        n_pop        = ga_cfg["n_pop"],
        G_max        = ga_cfg["G_max"],
        G_stag       = ga_cfg["G_stag"],
        k_tournament = ga_cfg["k_tournament"],
        delta_G      = alns_cfg["delta_G"],
        top_k_alns   = alns_cfg["top_k_alns"],
        time_limit_s = stop_cfg["time_limit_seconds"],
        milp_seed    = milp_seed,
        rng          = rng,
        log_callback = _log_cb,
    )
    return ga, alns


# ---------------------------------------------------------------------------
# Output builders
# ---------------------------------------------------------------------------

def extract_results(problem: Problem, sol, elapsed: float):
    p = problem
    schedule_rows = []
    cost_rows     = []
    grand = {k: 0.0 for k in ["overstock", "shortage", "backorder", "penalty_OA", "plt_penalty", "transport", "total"]}

    for t in p.periods:
        for wh in p.warehouses:
            inv   = sol.I.get((wh, t), 0.0)
            q_oa  = sol.Q_OA.get((wh, t), 0)
            lt_oa = p.LT_OA[wh] if isinstance(p.LT_OA, dict) else p.LT_OA
            q_received = sol.Q_OA.get((wh, t - lt_oa + 1), 0) if (t - lt_oa + 1) in p.periods else 0

            plt_in = sum(
                sol.Q_PLT.get((src, wh, t2), 0.0)
                for src in p.warehouses if src != wh
                for t2 in p.periods if t2 == t - p.LT_PLT.get((src, wh), 0)
            )
            plt_out = sum(sol.Q_PLT.get((wh, dst, t), 0.0) for dst in p.warehouses if dst != wh)

            u_val = p.U.get((wh, t), 1e9)
            l_val = p.L.get((wh, t), 0.0)
            Co_   = p.Co.get((wh, t), 0.0) * max(inv - u_val, 0.0)
            Cs_   = p.Cs.get((wh, t), 0.0) * max(l_val - inv, 0.0)
            Cb_   = p.Cb.get((wh, t), 0.0) * max(-inv, 0.0)
            r_val = sol.r_OA.get((wh, t), 0)
            Cp_   = p.Cp.get((wh, t), 0.0) if r_val > 0 else 0.0

            tc_row = sum(
                p.dist.get((i, j), 0.0) * p.TC
                for (i, j, t2), q in sol.Q_PLT.items()
                if t2 == t and j == wh and q > 0
            )

            # PLT partial-case penalty (per receiving wh at this period)
            plt_pen_row = sum(
                p.Cp_plt.get((src, wh, t), 0.0)
                for src in p.warehouses if src != wh
                if sol.Q_PLT.get((src, wh, t), 0.0) % max(p.case_pack(wh, t), 1) > 0
            )

            row_total = Co_ + Cs_ + Cb_ + Cp_ + plt_pen_row + tc_row

            inv_begin = sol.I.get((wh, t - 1), p.BI.get(wh, 0.0)) if t > p.periods[0] else p.BI.get(wh, 0.0)

            schedule_rows.append({
                "product"        : p.product,
                "week"           : t,
                "warehouse"      : wh,
                "inv_begin"      : round(inv_begin, 2),
                "Q_OA_allocated" : q_oa,
                "Q_OA_received"  : q_received,
                "PLT_in"         : round(plt_in, 2),
                "PLT_out"        : round(plt_out, 2),
                "delta_I"        : p.delta_I.get((wh, t), 0.0),
                "inventory_end"  : round(inv, 2),
                "surplus_E"      : round(max(inv - l_val, 0.0), 2),
                "backorder"      : round(max(-inv, 0.0), 2),
                "shortage"       : round(max(l_val - inv, 0.0), 2),
                "overstock"      : round(max(inv - u_val, 0.0), 2),
            })

            cost_rows.append({
                "product"           : p.product,
                "week"              : t,
                "warehouse"         : wh,
                "overstock_cost"    : round(Co_, 4),
                "shortage_cost"     : round(Cs_, 4),
                "backorder_cost"    : round(Cb_, 4),
                "penalty_OA_cost"   : round(Cp_, 4),
                "plt_penalty_cost"  : round(plt_pen_row, 4),
                "transport_cost"    : round(tc_row, 4),
                "total_cost"        : round(row_total, 4),
            })

            grand["overstock"]    += Co_
            grand["shortage"]     += Cs_
            grand["backorder"]    += Cb_
            grand["penalty_OA"]   += Cp_
            grand["plt_penalty"]  += plt_pen_row
            grand["transport"]    += tc_row
            grand["total"]        += row_total

    summary = {
        "product"                : p.product,
        "n_warehouses"           : p.n_wh,
        "n_periods"              : p.n_periods,
        "fitness"                : round(sol.fitness, 4),
        "elapsed_s"              : round(elapsed, 2),
        "total_overstock_cost"   : round(grand["overstock"],    4),
        "total_shortage_cost"    : round(grand["shortage"],     4),
        "total_backorder_cost"   : round(grand["backorder"],    4),
        "total_penalty_OA_cost"  : round(grand["penalty_OA"],  4),
        "total_plt_penalty_cost" : round(grand["plt_penalty"],  4),
        "total_transport_cost"   : round(grand["transport"],    4),
        "grand_total_cost"       : round(grand["total"],        4),
    }

    return schedule_rows, cost_rows, summary


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def write_csv(path: Path, rows: List[Dict], mode: str = "w") -> None:
    if not rows:
        return
    write_header = mode == "w" or not path.exists()
    with path.open(mode, newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        if write_header:
            w.writeheader()
        w.writerows(rows)


def get_writable_path(path: Path) -> Path:
    """If the file is locked by another process (e.g. Excel), append _1, _2... until writable."""
    if not path.exists():
        return path
    try:
        # Check write access
        with path.open("a", encoding="utf-8-sig") as f:
            pass
        return path
    except (IOError, PermissionError):
        suffix = 1
        while True:
            alt_path = path.parent / f"{path.stem}_{suffix}{path.suffix}"
            if not alt_path.exists():
                return alt_path
            try:
                with alt_path.open("a", encoding="utf-8-sig") as f:
                    pass
                return alt_path
            except (IOError, PermissionError):
                suffix += 1


# ---------------------------------------------------------------------------
# Excel pivot writer
# ---------------------------------------------------------------------------

def _num(v) -> float:
    """Safe cast to float for Excel cells."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


ROW_LABELS = [
    "Beginning inventory",
    "Inventory fluctuation",
    "Net inventory (before)",
    "OA Pack qty (arrived)",
    "OA Residual (arrived)",
    "PLT Pack qty (arrived)",
    "PLT Residual (arrived)",
    "PLT total sent out",
    "Net inventory (after)",
]


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _border() -> "Border":
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _wh_border() -> "Border":
    thick = Side(style="medium")
    thin  = Side(style="thin")
    return Border(left=thick, right=thick, top=thick, bottom=thick)


def build_pivot_data(problem: "Problem", sol) -> Dict:
    """
    For each warehouse, for each period, compute the 9 row values.
    Returns dict: wh -> row_label -> {t: value}
    """
    p   = problem
    out = {}

    for wh in p.warehouses:
        rows: Dict[str, Dict[int, float]] = {lbl: {} for lbl in ROW_LABELS}

        for t in p.periods:
            inv_begin = (
                sol.I.get((wh, t - 1), p.BI.get(wh, 0.0))
                if t > p.periods[0]
                else p.BI.get(wh, 0.0)
            )
            delta     = p.delta_I.get((wh, t), 0.0)

            # OA received this period = allocated LT_OA - 1 periods ago (running week is 0)
            lt_oa = p.LT_OA[wh] if isinstance(p.LT_OA, dict) else p.LT_OA
            q_oa_recv = sol.Q_OA.get((wh, t - lt_oa + 1), 0) if (t - lt_oa + 1) in p.periods else 0

            # PLT received this period (from all sources)
            plt_recv_total = sum(
                sol.Q_PLT.get((src, wh, t2), 0.0)
                for src in p.warehouses if src != wh
                for t2 in p.periods
                if t2 == t - p.LT_PLT.get((src, wh), 0)
            )

            # Net inventory before = I_{t-1} + OA_received + PLT_received + delta_I
            # This matches Decode Algorithm step 2: I^temp = I_{t-1} + Q^OA_{t-LT} + ΣQ^PLT_{t-LT_PLT} + ΔI
            net_before = inv_begin + q_oa_recv + plt_recv_total + delta
            
            cp_oa = p.case_pack(wh, t - lt_oa + 1)
            oa_pack = math.floor(q_oa_recv / cp_oa) * cp_oa if cp_oa > 0 else q_oa_recv
            oa_resid = q_oa_recv - oa_pack

            # PLT received already computed above for net_before
            cp_plt = p.case_pack(wh, t)
            plt_pack  = math.floor(plt_recv_total / cp_plt) * cp_plt if cp_plt > 0 else plt_recv_total
            plt_resid = plt_recv_total - plt_pack

            # PLT sent out from this warehouse this period
            plt_sent = sum(
                sol.Q_PLT.get((wh, dst, t), 0.0)
                for dst in p.warehouses if dst != wh
            )

            inv_after = sol.I.get((wh, t), 0.0)

            rows["Beginning inventory"][t]    = round(inv_begin, 1)
            rows["Inventory fluctuation"][t]  = round(delta, 1)
            rows["Net inventory (before)"][t] = round(net_before, 1)
            rows["OA Pack qty (arrived)"][t]  = round(oa_pack, 1)
            rows["OA Residual (arrived)"][t]  = round(oa_resid, 1)
            rows["PLT Pack qty (arrived)"][t] = round(plt_pack, 1)
            rows["PLT Residual (arrived)"][t] = round(plt_resid, 1)
            rows["PLT total sent out"][t]     = round(plt_sent, 1)
            rows["Net inventory (after)"][t]  = round(inv_after, 1)

        out[wh] = rows
    return out


def add_product_sheet(wb: "openpyxl.Workbook", problem: "Problem", sol, summary: Dict) -> None:
    """Add one worksheet per product in the pivot format shown in the image."""
    if not HAS_OPENPYXL:
        return

    p       = problem
    periods = list(p.periods)
    n_per   = len(periods)

    # Sheet name: product id (max 31 chars)
    sheet_name = p.product[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Color palette
    CLR_HEADER_DARK  = "2F5496"   # dark blue
    CLR_HEADER_MID   = "4472C4"   # medium blue
    CLR_WH_LABEL     = "D6E4F0"   # light blue
    CLR_NET_BEFORE   = "FFF2CC"   # yellow
    CLR_NET_AFTER_POS= "E2EFDA"   # green
    CLR_NET_AFTER_NEG= "FFDCE0"   # pink/red
    CLR_NET_AFTER_ZRO= "FFEB9C"   # orange-yellow
    CLR_OA_ROW       = "DDEEFF"   # light blue
    CLR_PLT_ROW      = "EAF4EA"   # light green
    CLR_WHITE        = "FFFFFF"
    CLR_GREY_LIGHT   = "F5F5F5"

    # ── Column layout ──────────────────────────────────────────────────────
    # Col A: Capacity label / Warehouse
    # Col B: Data Type
    # Col C: Case-pack
    # Col D..D+n_per-1: Period values
    COL_WH   = 1
    COL_TYPE = 2
    COL_CP   = 3
    COL_P1   = 4   # first period column

    def c(col): return get_column_letter(col)

    # Set column widths
    ws.column_dimensions[c(COL_WH)].width   = 12
    ws.column_dimensions[c(COL_TYPE)].width = 26
    ws.column_dimensions[c(COL_CP)].width   = 10
    for i in range(n_per):
        ws.column_dimensions[c(COL_P1 + i)].width = 8

    # ── Row 1: capacity header ─────────────────────────────────────────────
    row = 1
    # Capacity from CAP dict (take period 1 value or max)
    cap_val = p.CAP.get(periods[0], 0) if p.CAP else 0
    cap_cell = ws.cell(row=row, column=COL_WH,
                       value=f"Capacity = {int(cap_val)}  |  Product: {p.product}  |  CP={p.CP}")
    cap_cell.font = Font(bold=True, color="FFFFFF")
    cap_cell.fill = _fill(CLR_HEADER_DARK)
    ws.merge_cells(start_row=row, start_column=COL_WH,
                   end_row=row, end_column=COL_CP)

    period_lbl = ws.cell(row=row, column=COL_P1, value="Period")
    period_lbl.font = Font(bold=True, color="FFFFFF")
    period_lbl.fill = _fill(CLR_HEADER_DARK)
    period_lbl.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=COL_P1,
                   end_row=row, end_column=COL_P1 + n_per - 1)

    # ── Row 2: sub-header ─────────────────────────────────────────────────
    row = 2
    for col, txt in [(COL_WH, "Warehouse"), (COL_TYPE, "Data Type"), (COL_CP, "Case-pack")]:
        cell = ws.cell(row=row, column=col, value=txt)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill(CLR_HEADER_MID)
        cell.alignment = Alignment(horizontal="center")
    for i, t in enumerate(periods):
        cell = ws.cell(row=row, column=COL_P1 + i, value=t)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill(CLR_HEADER_MID)
        cell.alignment = Alignment(horizontal="center")

    # ── Build pivot data ──────────────────────────────────────────────────
    pivot = build_pivot_data(problem, sol)

    # ── Write warehouse blocks ────────────────────────────────────────────
    current_row = 3

    for wh in p.warehouses:
        wh_data    = pivot[wh]
        n_rows_wh  = len(ROW_LABELS)
        wh_start   = current_row
        wh_end     = current_row + n_rows_wh - 1

        u_val = p.U.get((wh, periods[0]), 1e9)   # use first period ceiling as reference
        l_val = p.L.get((wh, periods[0]), 0.0)

        for ri, label in enumerate(ROW_LABELS):
            r = current_row + ri

            # Warehouse label (only first row of wh block)
            wh_cell = ws.cell(row=r, column=COL_WH)
            if ri == 0:
                wh_cell.value = wh
                wh_cell.font  = Font(bold=True)
            wh_cell.fill = _fill(CLR_WH_LABEL)

            # Data type label
            type_cell = ws.cell(row=r, column=COL_TYPE, value=label)
            type_cell.font = Font(italic=(label not in (
                "Net inventory (before)", "Net inventory (after)"
            )))

            # Case-pack column
            cp_cell = ws.cell(row=r, column=COL_CP, value=p.case_pack(wh, periods[0]))
            cp_cell.alignment = Alignment(horizontal="center")

            # Period values
            for ci, t in enumerate(periods):
                val  = wh_data[label].get(t, 0.0)
                cell = ws.cell(row=r, column=COL_P1 + ci, value=val)
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.##"

                # Row background
                if label == "Net inventory (before)":
                    cell.fill = _fill(CLR_NET_BEFORE)
                    type_cell.fill = _fill(CLR_NET_BEFORE)
                elif label == "Net inventory (after)":
                    if val > 0:
                        cell.fill = _fill(CLR_NET_AFTER_POS)
                    elif val < 0:
                        cell.fill = _fill(CLR_NET_AFTER_NEG)
                    else:
                        cell.fill = _fill(CLR_NET_AFTER_ZRO)
                    type_cell.fill = _fill(CLR_NET_AFTER_POS if val > 0
                                           else (CLR_NET_AFTER_NEG if val < 0 else CLR_NET_AFTER_ZRO))
                    cell.font = Font(bold=True)
                elif "OA" in label:
                    cell.fill = _fill(CLR_OA_ROW)
                elif "PLT" in label:
                    cell.fill = _fill(CLR_PLT_ROW)
                else:
                    cell.fill = _fill(CLR_WHITE if ri % 2 == 0 else CLR_GREY_LIGHT)

        # Merge warehouse label cells vertically
        if n_rows_wh > 1:
            ws.merge_cells(start_row=wh_start, start_column=COL_WH,
                           end_row=wh_end,   end_column=COL_WH)
            wh_cell_merged = ws.cell(row=wh_start, column=COL_WH)
            wh_cell_merged.alignment = Alignment(vertical="center", horizontal="center", wrap_text=False)

        # Thick border around warehouse block
        thick = Side(style="medium")
        thin  = Side(style="thin")
        for ri2 in range(n_rows_wh):
            r2 = wh_start + ri2
            for col2 in range(COL_WH, COL_P1 + n_per):
                cell2 = ws.cell(row=r2, column=col2)
                top_s    = thick if ri2 == 0            else thin
                bottom_s = thick if ri2 == n_rows_wh-1 else thin
                left_s   = thick if col2 == COL_WH     else thin
                right_s  = thick if col2 == COL_P1+n_per-1 else thin
                cell2.border = Border(left=left_s, right=right_s,
                                      top=top_s, bottom=bottom_s)

        current_row = wh_end + 1

    # ── Cost summary rows at bottom ───────────────────────────────────────
    current_row += 1
    labels_costs = [
        ("Grand Total Cost",    summary["grand_total_cost"]),
        ("  Overstock",         summary["total_overstock_cost"]),
        ("  Shortage",          summary["total_shortage_cost"]),
        ("  Backorder",         summary["total_backorder_cost"]),
        ("  Penalty OA",        summary["total_penalty_OA_cost"]),
        ("  PLT Penalty",       summary.get("total_plt_penalty_cost", 0.0)),
        ("  Transport",         summary["total_transport_cost"]),
        ("Fitness",             summary["fitness"]),
        ("Elapsed (s)",         summary["elapsed_s"]),
    ]
    for lbl, val in labels_costs:
        ws.cell(row=current_row, column=COL_WH, value="").fill  = _fill("F0F0F0")
        ws.cell(row=current_row, column=COL_TYPE, value=lbl).font = Font(bold=lbl.startswith("G"))
        ws.cell(row=current_row, column=COL_P1, value=val).number_format = "#,##0.00"
        current_row += 1

    # Freeze header rows
    ws.freeze_panes = ws.cell(row=3, column=COL_P1)


def save_excel(wb: "openpyxl.Workbook", path: Path) -> None:
    if not HAS_OPENPYXL:
        return
    # Ensure the default empty sheet is removed
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(str(path))
    log.info("Saved Excel: %s", path)


# ---------------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------------

def main() -> None:
    global DATA_DIR
    parser = argparse.ArgumentParser(description="Run GA-ALNS for products in inventory_flow.csv")
    parser.add_argument("--data-dir", type=str, default=str(BASE_DIR / "data"), help="Path to data directory")
    parser.add_argument("--max-products", type=int, default=999999, help="Max products to run (default all)")
    parser.add_argument("--config", type=str, default=str(BASE_DIR / "model" / "config.json"))
    parser.add_argument("--no-excel", action="store_true", help="Skip Excel output")
    args = parser.parse_args()

    config_path = Path(args.config)
    if not config_path.exists():
        log.error("Config not found: %s", config_path)
        sys.exit(1)

    DATA_DIR = Path(args.data_dir)
    cfg  = load_config(config_path)
    seed = cfg.get("seed", 42)

    # Load shared data tables
    log.info("Loading CSV data...")
    product_ids, flow_data = load_inventory_flow(n=args.max_products)
    bi_data     = load_inventory_begin()
    costs_data  = load_unit_costs()
    cap_data    = load_vendor_capacity()
    cp_data     = load_packing_details()
    lt_plt      = load_plt_lead_times()
    dist_data   = load_distances()
    LT_OA       = load_oa_lead_times()
    TC          = 1.2  # container 40ft Dry (container_pricing.csv id=2)

    log.info("Found %d products to process.", len(product_ids))

    # Prepare output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    schedule_path = get_writable_path(OUTPUT_DIR / "schedule.csv")
    costs_path    = get_writable_path(OUTPUT_DIR / "costs.csv")
    summary_path  = get_writable_path(OUTPUT_DIR / "cost_summary.csv")
    log_path      = get_writable_path(OUTPUT_DIR / "run_log.csv")
    excel_path    = get_writable_path(OUTPUT_DIR / "schedule_detail.xlsx")

    # Clear existing files (if they are writable)
    for p in [schedule_path, costs_path, summary_path, log_path, excel_path]:
        try:
            if p.exists():
                p.unlink()
        except Exception as e:
            log.warning("Could not clear existing file %s: %s", p.name, e)

    # Prepare Excel workbook
    write_excel = HAS_OPENPYXL and not args.no_excel
    if write_excel:
        wb = openpyxl.Workbook()
    else:
        wb = None
        if not HAS_OPENPYXL:
            log.warning("openpyxl not installed — Excel output skipped. Run: pip install openpyxl")

    all_summaries: List[Dict] = []
    success = 0
    failed  = 0

    for idx, pid in enumerate(product_ids, 1):
        log.info("=" * 60)
        log.info("[%d/%d] Product: %s", idx, len(product_ids), pid)
        log.info("=" * 60)

        try:
            # Build problem from CSV data
            pflow   = flow_data.get(pid, {})
            bi_map  = bi_data.get(pid, {})
            c_map   = costs_data.get(pid, {})
            cap_map = cap_data.get(pid, {})
            cp      = cp_data.get(pid, 1)

            if not pflow:
                log.warning("  No inventory_flow data for %s — skipping.", pid)
                failed += 1
                continue

            if not bi_map:
                log.warning("  No inventory_begin data for %s — skipping.", pid)
                failed += 1
                continue

            problem = build_problem(
                product_id = pid,
                flow       = pflow,
                bi_map     = bi_map,
                costs_map  = c_map,
                cap_map    = cap_map,
                cp         = cp,
                lt_oa      = LT_OA,
                lt_plt_map = lt_plt,
                dist_map   = dist_data,
                tc         = TC,
            )

            log.info("  %d warehouses | %d periods | CP=%d", problem.n_wh, problem.n_periods, problem.CP)

            rng = random.Random(seed + idx)
            ga, _ = build_components(problem, cfg, rng)

            t0 = time.perf_counter()
            _, best_sol = ga.run()
            elapsed = time.perf_counter() - t0

            log.info("  Done: fitness=%.4f | elapsed=%.1fs", best_sol.fitness, elapsed)

            sched_rows, cost_rows, summary = extract_results(problem, best_sol, elapsed)

            # Append to aggregate files
            write_csv(schedule_path, sched_rows, mode="a")
            write_csv(costs_path,    cost_rows,  mode="a")
            all_summaries.append(summary)

            # Add Excel pivot sheet for this product
            if write_excel:
                add_product_sheet(wb, problem, best_sol, summary)
                if idx % 10 == 0 or idx == len(product_ids):
                    save_excel(wb, excel_path)   # checkpoint save

            # Log to console
            log.info("  Overstock : %12.2f", summary["total_overstock_cost"])
            log.info("  Shortage  : %12.2f", summary["total_shortage_cost"])
            log.info("  Backorder : %12.2f", summary["total_backorder_cost"])
            log.info("  Penalty OA: %12.2f", summary["total_penalty_OA_cost"])
            log.info("  Transport : %12.2f", summary["total_transport_cost"])
            log.info("  TOTAL     : %12.2f", summary["grand_total_cost"])

            success += 1
        except Exception as e:
            log.error("  FAILED for %s: %s", pid, e)
            log.debug(traceback.format_exc())
            failed += 1
            # Write a failed row to summary
            all_summaries.append({
                "product"                : pid,
                "n_warehouses"           : 0,
                "n_periods"              : 0,
                "fitness"                : None,
                "elapsed_s"              : None,
                "total_overstock_cost"   : None,
                "total_shortage_cost"    : None,
                "total_backorder_cost"   : None,
                "total_penalty_OA_cost"  : None,
                "total_plt_penalty_cost" : None,
                "total_transport_cost"   : None,
                "grand_total_cost"       : None,
            })

    # Write cost_summary.csv
    if all_summaries:
        write_csv(summary_path, all_summaries, mode="w")
        log.info("Saved cost summary: %s", summary_path)

    # Write run_log.csv (subset of summary)
    log_rows = []
    for s in all_summaries:
        log_rows.append({
            "product"        : s["product"],
            "n_warehouses"   : s["n_warehouses"],
            "n_periods"      : s["n_periods"],
            "fitness"        : s["fitness"],
            "elapsed_s"      : s["elapsed_s"],
            "grand_total_cost": s["grand_total_cost"],
            "status"         : "ok" if s["fitness"] is not None else "failed",
        })
    write_csv(log_path, log_rows, mode="w")

    # Final Excel save
    if write_excel:
        save_excel(wb, excel_path)

    log.info("=" * 60)
    log.info("DONE: %d success, %d failed", success, failed)
    log.info("Outputs in: %s", OUTPUT_DIR)
    log.info("  schedule.csv         — full schedule (flat CSV)")
    log.info("  costs.csv            — per-row cost breakdown")
    log.info("  cost_summary.csv     — total + component costs per product")
    log.info("  run_log.csv          — timing and fitness summary")
    if write_excel:
        log.info("  schedule_detail.xlsx — pivot schedule, 1 sheet/product")


if __name__ == "__main__":
    main()
