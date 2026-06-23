from __future__ import annotations

import argparse
import csv
import logging
import sys
import time
import zipfile
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    import gurobipy as gp
    from gurobipy import GRB
except ImportError:
    gp = None
    GRB = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# ============================================================
# Logging
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)

log = logging.getLogger("milp_new_model")


# ============================================================
# Paths
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "sample_100_fixed_old"
OUTPUT_DIR = BASE_DIR / "output_milp_new_model_1"


# ============================================================
# Helpers
# ============================================================

def _format_wh(val: str) -> str:
    val = str(val).strip()

    if val.startswith("WH"):
        return val

    try:
        return f"WH{int(val):02d}"
    except ValueError:
        return val


def _case_pack(problem: Dict[str, Any], wh: str, t: int) -> int:
    cp = problem["CP"]

    if isinstance(cp, dict):
        value = cp.get((wh, t)) or cp.get(wh) or cp.get(t) or 1
    else:
        value = cp

    return max(int(round(float(value))), 1)


def _lt_oa(problem: Dict[str, Any], wh: str) -> int:
    lt = problem["LT_OA"]

    if isinstance(lt, dict):
        return int(lt.get(wh, 0))

    return int(lt)


def _oa_order_period(
    problem: Dict[str, Any],
    wh: str,
    arrival_t: int,
    arrival_indexing: str,
) -> int:
    """
    Return the OA decision period whose shipment arrives at arrival_t.

    business mode:
        order at period 1 with LT=8 arrives at period 8,
        so order_t = t - LT + 1.

    math mode:
        follows compact formula Q_{t-LT},
        so order_t = t - LT.
    """
    lt = _lt_oa(problem, wh)

    if arrival_indexing == "math":
        return arrival_t - lt

    return arrival_t - lt + 1


def _build_big_m(problem: Dict[str, Any]) -> float:
    periods = problem["periods"]
    warehouses = problem["warehouses"]

    total_cap = sum(abs(problem["CAP"].get(t, 0.0)) for t in periods)

    total_delta = sum(
        abs(problem["delta_I"].get((i, t), 0.0))
        for i in warehouses
        for t in periods
    )

    max_bi = max([abs(problem["BI"].get(i, 0.0)) for i in warehouses] + [0.0])

    max_bound = max(
        [
            max(
                abs(problem["L"].get((i, t), 0.0)),
                abs(problem["U"].get((i, t), 0.0)),
            )
            for i in warehouses
            for t in periods
        ]
        + [0.0]
    )

    return max(
        100_000.0,
        10.0 * (total_cap + total_delta + max_bi + max_bound + 1.0),
    )


def _excel_value(value: Any) -> Any:
    if value is None:
        return None

    try:
        x = float(value)

        if abs(x - round(x)) < 1e-6:
            return int(round(x))

        return round(x, 4)
    except Exception:
        return value


def _safe_sheet_name(name: str, used_names: set[str]) -> str:
    invalid = ["\\", "/", "*", "?", ":", "[", "]"]

    for ch in invalid:
        name = name.replace(ch, "_")

    name = name[:31]
    base = name
    counter = 1

    while name in used_names:
        suffix = f"_{counter}"
        name = base[:31 - len(suffix)] + suffix
        counter += 1

    used_names.add(name)
    return name


def _status_name(model) -> str:
    status_map = {
        GRB.OPTIMAL: "optimal",
        GRB.TIME_LIMIT: "time_limit",
        GRB.INFEASIBLE: "infeasible",
        GRB.INF_OR_UNBD: "inf_or_unbd",
        GRB.UNBOUNDED: "unbounded",
        GRB.INTERRUPTED: "interrupted",
    }

    mem_limit_code = getattr(GRB, "MEM_LIMIT", None)

    if mem_limit_code is not None:
        status_map[mem_limit_code] = "mem_limit"

    return status_map.get(model.Status, str(model.Status))


def _empty_summary(
    product: str,
    status: str,
    elapsed: float,
    error_message: str | None = None,
    model=None,
) -> Dict[str, Any]:
    return {
        "product": product,
        "solver_status": status,
        "objective_value": None,
        "recomputed_total_cost": None,
        "best_bound": None,
        "mip_gap": None,
        "runtime_seconds": round(elapsed, 4),
        "num_variables": getattr(model, "NumVars", None) if model is not None else None,
        "num_binary_variables": getattr(model, "NumBinVars", None) if model is not None else None,
        "num_integer_variables": getattr(model, "NumIntVars", None) if model is not None else None,
        "num_constraints": getattr(model, "NumConstrs", None) if model is not None else None,
        "node_count": round(getattr(model, "NodeCount", 0), 0) if model is not None else None,
        "simplex_iterations": round(getattr(model, "IterCount", 0), 0) if model is not None else None,
        "error_message": error_message,
        "total_overstock_cost": None,
        "total_shortage_cost": None,
        "total_backorder_cost": None,
        "total_penalty_OA_cost": None,
        "total_plt_penalty_cost": None,
        "total_transport_cost": None,
        "grand_total_cost": None,
    }


def _failed_result(
    product: str,
    status: str,
    elapsed: float,
    error_message: str | None = None,
    model=None,
) -> Dict[str, Any]:
    return {
        "product": product,
        "status": status,
        "objective_value": None,
        "best_bound": None,
        "mip_gap": None,
        "runtime_seconds": elapsed,
        "schedule_rows": [],
        "cost_rows": [],
        "plt_rows": [],
        "detail_sheet_rows": [],
        "summary": _empty_summary(product, status, elapsed, error_message, model),
    }


def _maybe_extract_zip(path: Path) -> Tuple[Path, tempfile.TemporaryDirectory | None]:
    if path.is_file() and path.suffix.lower() == ".zip":
        tmp = tempfile.TemporaryDirectory()

        with zipfile.ZipFile(path, "r") as zf:
            zf.extractall(tmp.name)

        extracted = Path(tmp.name)
        csv_names = {"inventory_flow.csv", "inventory_begin.csv", "unit_cost.csv"}

        if csv_names.issubset({p.name for p in extracted.iterdir()}):
            return extracted, tmp

        for child in extracted.iterdir():
            if child.is_dir() and csv_names.issubset({p.name for p in child.iterdir()}):
                return child, tmp

        return extracted, tmp

    return path, None


# ============================================================
# Data loaders
# ============================================================

def load_inventory_flow(n: int = 999999) -> Tuple[List[str], Dict]:
    path = DATA_DIR / "inventory_flow.csv"

    product_order: List[str] = []
    flow: Dict[str, Dict[str, Dict[int, Dict]]] = defaultdict(lambda: defaultdict(dict))

    with path.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            pid = row["product_id"]
            wh = _format_wh(row["warehouse_id"])
            t = int(row["time_period"])

            if pid not in product_order:
                if len(product_order) >= n:
                    continue
                product_order.append(pid)

            if pid in product_order:
                flow[pid][wh][t] = {
                    "delta_I": float(row["inventory_fluctuation"]),
                    "U": float(row["inventory_ceiling"]),
                    "L": float(row["inventory_floor"]),
                }

    return product_order, flow


def load_inventory_begin() -> Dict[str, Dict[str, float]]:
    bi: Dict[str, Dict[str, float]] = defaultdict(dict)

    with (DATA_DIR / "inventory_begin.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            bi[row["product_id"]][_format_wh(row["warehouse_id"])] = float(
                row["beginning_inventory"]
            )

    return bi


def load_unit_costs() -> Dict[str, Dict[str, Dict[int, Dict]]]:
    costs: Dict[str, Dict[str, Dict[int, Dict]]] = defaultdict(lambda: defaultdict(dict))

    with (DATA_DIR / "unit_cost.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            pid = row["product_id"]
            wh = _format_wh(row["warehouse_id"])
            t = int(row["time_period"])

            costs[pid][wh][t] = {
                "Co": float(row["overstock_cost"]),
                "Cs": float(row["shortage_cost"]),
                "Cb": float(row["backlog_cost"]),
                "Cp": float(row["penalty_cost"]),
            }

    return costs


def load_vendor_capacity() -> Dict[str, Dict[int, float]]:
    cap: Dict[str, Dict[int, float]] = defaultdict(dict)

    with (DATA_DIR / "vendor_capacity.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cap[row["product_id"]][int(row["time_period"])] = float(row["capacity"])

    return cap


def load_packing_details() -> Dict[str, int]:
    cp: Dict[str, int] = {}

    with (DATA_DIR / "packing_details.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cp[row["product_id"]] = int(float(row["pack_multiple"]))

    return cp


def load_oa_lead_times() -> Dict[str, int]:
    path = DATA_DIR / "oa_lead_time.csv"
    lt: Dict[str, int] = {}
    
    if path.exists():
        with path.open(encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                wh = _format_wh(row["warehouse_id"])
                lt[wh] = int(float(row["week_lead_time"]))
    
    return lt


def load_plt_lead_times() -> Dict[Tuple[str, str], int]:
    lt: Dict[Tuple[str, str], int] = {}

    with (DATA_DIR / "plt_lead_time.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            frm = _format_wh(row["from_warehouse_id"])
            to = _format_wh(row["to_warehouse_id"])
            lt[(frm, to)] = int(float(row["lead_time_weeks"]))

    return lt


def load_distances() -> Dict[Tuple[str, str], float]:
    state_map = {
        "MI": "WH01",
        "OH": "WH02",
        "IN": "WH03",
        "IL": "WH04",
        "KY": "WH05",
        "MO": "WH06",
    }

    dist: Dict[Tuple[str, str], float] = {}

    with (DATA_DIR / "FGPs_distance.csv").open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            from_state = row["State"].strip()
            from_wh = state_map.get(from_state)

            if not from_wh:
                continue

            for state, wh in state_map.items():
                if state in row and row[state] not in (None, ""):
                    dist[(from_wh, wh)] = float(row[state])

    return dist


# ============================================================
# Problem builder
# ============================================================

def build_problem(
    product_id: str,
    flow: Dict[str, Dict[int, Dict]],
    bi_map: Dict[str, float],
    costs_map: Dict[str, Dict[int, Dict]],
    cap_map: Dict[int, float],
    cp: int,
    lt_oa: int | Dict[str, int],
    lt_plt_map: Dict[Tuple[str, str], int],
    dist_map: Dict[Tuple[str, str], float],
    tc: float = 1.2,
) -> Dict[str, Any]:
    warehouses = tuple(sorted(flow.keys()))

    periods_set = set()

    for wh_data in flow.values():
        periods_set.update(wh_data.keys())

    periods = tuple(sorted(periods_set))

    LT_PLT = {
        (i, j): int(v)
        for (i, j), v in lt_plt_map.items()
        if i in warehouses and j in warehouses and i != j
    }

    if isinstance(lt_oa, dict):
        LT_OA = {wh: int(lt_oa.get(wh, 8)) for wh in warehouses}
    else:
        LT_OA = {wh: int(lt_oa) for wh in warehouses}

    PLT_periods = {
        wh: frozenset(t for t in periods if t <= LT_OA[wh] - 1)
        for wh in warehouses
    }

    BI = {wh: bi_map.get(wh, 0.0) for wh in warehouses}

    delta_I = {}
    U = {}
    L = {}

    for wh in warehouses:
        for t in periods:
            entry = flow[wh].get(t, {})
            delta_I[(wh, t)] = entry.get("delta_I", 0.0)
            U[(wh, t)] = entry.get("U", 1e6)
            L[(wh, t)] = entry.get("L", 0.0)

    CAP = {t: cap_map.get(t, 9999.0) for t in periods}

    Co = {}
    Cs = {}
    Cb = {}
    Cp_map = {}

    for wh in warehouses:
        for t in periods:
            c = costs_map.get(wh, {}).get(t, {})
            Co[(wh, t)] = c.get("Co", 0.1)
            Cs[(wh, t)] = c.get("Cs", 0.5)
            Cb[(wh, t)] = c.get("Cb", 1500.0)
            Cp_map[(wh, t)] = c.get("Cp", 2000.0)

    dist = {
        (i, j): dist_map.get((i, j), 0.0)
        for i in warehouses
        for j in warehouses
    }

    Cp_plt = {}

    for (i, j) in LT_PLT:
        for t in periods:
            Cp_plt[(i, j, t)] = Cp_map.get((j, t), 0.0)

    return {
        "product": product_id,
        "warehouses": warehouses,
        "periods": periods,
        "LT_OA": LT_OA,
        "LT_PLT": LT_PLT,
        "PLT_periods": PLT_periods,
        "BI": BI,
        "delta_I": delta_I,
        "U": U,
        "L": L,
        "CAP": CAP,
        "CP": cp,
        "TC": tc,
        "dist": dist,
        "Co": Co,
        "Cs": Cs,
        "Cb": Cb,
        "Cp": Cp_map,
        "Cp_plt": Cp_plt,
    }


# ============================================================
# Excel detail builder
# ============================================================

def build_detail_sheet_rows_from_solution(
    problem: Dict[str, Any],
    T: List[int],
    W: List[str],
    arcs: List[Tuple[str, str, int]],
    Q_OA,
    q_OA,
    r_OA,
    Q_PLT,
    q_PLT,
    r_PLT,
    I_pre,
    I,
    summary: Dict[str, Any],
    val,
    arrival_indexing: str,
) -> List[List[Any]]:
    product = problem["product"]

    caps = [int(round(problem["CAP"].get(t, 0.0))) for t in T]
    cap_text = (
        f"Capacity = {caps[0]}"
        if len(set(caps)) == 1
        else f"Capacity = {min(caps)}--{max(caps)}"
    )

    cp_value = problem["CP"]
    cp_text = cp_value if not isinstance(cp_value, dict) else "varies"

    n_cols = 3 + len(T)
    rows: List[List[Any]] = []

    title_row = [""] * n_cols
    title_row[0] = f"{cap_text} | Product: {product} | CP = {cp_text}"
    title_row[3] = "Period"
    rows.append(title_row)

    rows.append(["Warehouse", "Data type", "Case-pack"] + T)

    def oa_pack_arrived(i: str, t: int) -> float:
        order_t = _oa_order_period(problem, i, t, arrival_indexing)

        if order_t not in T:
            return 0.0

        cp = _case_pack(problem, i, order_t)
        return val(q_OA[i, order_t]) * cp

    def oa_residual_arrived(i: str, t: int) -> float:
        order_t = _oa_order_period(problem, i, t, arrival_indexing)

        if order_t not in T:
            return 0.0

        return val(r_OA[i, order_t])

    def plt_pack_arrived(i: str, t: int) -> float:
        total = 0.0

        for src, dst, ship_t in arcs:
            if dst != i:
                continue

            lt = int(problem["LT_PLT"][(src, dst)])

            if ship_t + lt == t:
                cp = _case_pack(problem, dst, ship_t)
                total += val(q_PLT[src, dst, ship_t]) * cp

        return total

    def plt_residual_arrived(i: str, t: int) -> float:
        total = 0.0

        for src, dst, ship_t in arcs:
            if dst != i:
                continue

            lt = int(problem["LT_PLT"][(src, dst)])

            if ship_t + lt == t:
                total += val(r_PLT[src, dst, ship_t])

        return total

    def plt_total_sent_out(i: str, t: int) -> float:
        return sum(
            val(Q_PLT[src, dst, tt])
            for src, dst, tt in arcs
            if src == i and tt == t
        )

    for wh in W:
        cp = _case_pack(problem, wh, T[0])

        beginning_inventory = []
        inventory_fluctuation = []
        net_inventory_before = []
        oa_pack = []
        oa_residual = []
        plt_pack = []
        plt_residual = []
        plt_out = []
        net_inventory_after = []

        for idx, t in enumerate(T):
            begin_inv = problem["BI"].get(wh, 0.0) if idx == 0 else val(I[wh, T[idx - 1]])

            beginning_inventory.append(_excel_value(begin_inv))
            inventory_fluctuation.append(_excel_value(problem["delta_I"].get((wh, t), 0.0)))
            net_inventory_before.append(_excel_value(val(I_pre[wh, t])))
            oa_pack.append(_excel_value(oa_pack_arrived(wh, t)))
            oa_residual.append(_excel_value(oa_residual_arrived(wh, t)))
            plt_pack.append(_excel_value(plt_pack_arrived(wh, t)))
            plt_residual.append(_excel_value(plt_residual_arrived(wh, t)))
            plt_out.append(_excel_value(plt_total_sent_out(wh, t)))
            net_inventory_after.append(_excel_value(val(I[wh, t])))

        rows.append([wh, "Beginning inventory", cp] + beginning_inventory)
        rows.append(["", "Inventory fluctuation", cp] + inventory_fluctuation)
        rows.append(["", "Net inventory before PLT", cp] + net_inventory_before)
        rows.append(["", "OA pack arrived", cp] + oa_pack)
        rows.append(["", "OA residual arrived", cp] + oa_residual)
        rows.append(["", "PLT pack arrived", cp] + plt_pack)
        rows.append(["", "PLT residual arrived", cp] + plt_residual)
        rows.append(["", "PLT sent out", cp] + plt_out)
        rows.append(["", "Net inventory after PLT", cp] + net_inventory_after)

    rows.append([""] * n_cols)

    summary_items = [
        ("Grand total cost", "grand_total_cost"),
        ("  Overstock cost", "total_overstock_cost"),
        ("  Shortage cost", "total_shortage_cost"),
        ("  Backorder cost", "total_backorder_cost"),
        ("  OA penalty cost", "total_penalty_OA_cost"),
        ("  PLT penalty cost", "total_plt_penalty_cost"),
        ("  Transport cost", "total_transport_cost"),
        ("Fitness", "objective_value"),
        ("Runtime seconds", "runtime_seconds"),
        ("Solver status", "solver_status"),
        ("MIP gap", "mip_gap"),
        ("Node count", "node_count"),
        ("Variables", "num_variables"),
        ("Binary variables", "num_binary_variables"),
        ("Integer variables", "num_integer_variables"),
        ("Constraints", "num_constraints"),
        ("Error message", "error_message"),
    ]

    for label, key in summary_items:
        row = [""] * n_cols
        row[1] = label
        row[3] = summary.get(key)
        rows.append(row)

    return rows


def write_schedule_detail_xlsx(
    detail_sheets: Dict[str, List[List[Any]]],
    output_path: Path,
) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Run: python -m pip install openpyxl")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_names: set[str] = set()

    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    warehouse_fill = PatternFill("solid", fgColor="E2F0D9")
    summary_fill = PatternFill("solid", fgColor="FFF2CC")

    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    summary_labels = {
        "Grand total cost",
        "  Overstock cost",
        "  Shortage cost",
        "  Backorder cost",
        "  OA penalty cost",
        "  PLT penalty cost",
        "  Transport cost",
        "Fitness",
        "Runtime seconds",
        "Solver status",
        "MIP gap",
        "Node count",
        "Variables",
        "Binary variables",
        "Integer variables",
        "Constraints",
        "Error message",
    }

    for product_id, rows in detail_sheets.items():
        ws = wb.create_sheet(_safe_sheet_name(str(product_id), used_names))

        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                cell.alignment = center

                if c_idx == 2:
                    cell.alignment = left

                if isinstance(value, float):
                    cell.number_format = "#,##0.0000"
                elif isinstance(value, int):
                    cell.number_format = "#,##0"

        max_col = max(len(row) for row in rows)

        if max_col >= 4:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=max_col)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center

        for cell in ws[2]:
            cell.fill = subheader_fill
            cell.font = bold_font
            cell.alignment = center

        for row_idx in range(3, len(rows) + 1):
            label_a = ws.cell(row=row_idx, column=1).value
            label_b = ws.cell(row=row_idx, column=2).value

            if label_a:
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = warehouse_fill
                    ws.cell(row=row_idx, column=col_idx).font = bold_font

            if label_b in summary_labels:
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = summary_fill

                ws.cell(row=row_idx, column=2).font = bold_font
                ws.cell(row=row_idx, column=4).font = bold_font

        ws.freeze_panes = "D3"
        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 12

        for col_idx in range(4, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        for row_idx in range(1, len(rows) + 1):
            ws.row_dimensions[row_idx].height = 22

        ws.sheet_view.showGridLines = False

    wb.save(output_path)


# ============================================================
# MILP solver for the new mathematical model
# ============================================================

def solve_milp_new_model(
    problem: Dict[str, Any],
    time_limit: float | None = None,
    mip_gap: float = 0.0,
    threads: int = 0,
    verbose: bool = False,
    arrival_indexing: str = "business",
) -> Dict[str, Any]:
    if gp is None:
        raise RuntimeError("gurobipy is not installed. Run: python -m pip install gurobipy")

    product = problem["product"]
    W = list(problem["warehouses"])
    T = list(problem["periods"])
    M = _build_big_m(problem)

    start = time.perf_counter()
    model = None

    try:
        arcs = []

        for i in W:
            for j in W:
                if i == j:
                    continue

                if (i, j) not in problem["LT_PLT"]:
                    continue

                for t in T:
                    if t in problem["PLT_periods"].get(j, frozenset()):
                        arcs.append((i, j, t))

        model = gp.Model(f"MILP_NEW_{product}")

        nodefile_dir = OUTPUT_DIR / "gurobi_nodefiles"
        nodefile_dir.mkdir(parents=True, exist_ok=True)

        model.Params.OutputFlag = 1 if verbose else 0
        model.Params.MIPGap = mip_gap
        model.Params.Threads = threads if threads > 0 else 2
        model.Params.NodefileStart = 0.5
        model.Params.NodefileDir = str(nodefile_dir)
        model.Params.MIPFocus = 1

        if time_limit is not None and time_limit > 0:
            model.Params.TimeLimit = time_limit

        # ====================================================
        # Decision variables
        # ====================================================

        Q_OA = model.addVars(W, T, vtype=GRB.INTEGER, lb=0, name="Q_OA")
        q_OA = model.addVars(W, T, vtype=GRB.INTEGER, lb=0, name="q_OA")
        r_OA = model.addVars(W, T, vtype=GRB.INTEGER, lb=0, name="r_OA")
        y_OA = model.addVars(W, T, vtype=GRB.BINARY, name="y_OA")

        Q_PLT = model.addVars(arcs, vtype=GRB.INTEGER, lb=0, name="Q_PLT")
        q_PLT = model.addVars(arcs, vtype=GRB.INTEGER, lb=0, name="q_PLT")
        r_PLT = model.addVars(arcs, vtype=GRB.INTEGER, lb=0, name="r_PLT")
        y_PLT = model.addVars(arcs, vtype=GRB.BINARY, name="y_PLT")
        z_PLT = model.addVars(arcs, vtype=GRB.BINARY, name="z_PLT")

        I_pre = model.addVars(W, T, vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, name="I_pre")
        I = model.addVars(W, T, vtype=GRB.CONTINUOUS, lb=-GRB.INFINITY, name="I")

        available = model.addVars(W, T, vtype=GRB.CONTINUOUS, lb=0, name="available_after_floor")
        available_flag = model.addVars(W, T, vtype=GRB.BINARY, name="available_flag")

        overstock = model.addVars(W, T, vtype=GRB.CONTINUOUS, lb=0, name="overstock")
        shortage = model.addVars(W, T, vtype=GRB.CONTINUOUS, lb=0, name="shortage")
        backorder = model.addVars(W, T, vtype=GRB.CONTINUOUS, lb=0, name="backorder")

        # ====================================================
        # OA capacity and case-pack constraints
        # ====================================================

        for t in T:
            lhs = gp.quicksum(Q_OA[i, t] for i in W)
            cap = int(round(problem["CAP"].get(t, 0.0)))

            model.addConstr(
                lhs == cap,
                name=f"capacity_eq_{t}",
            )

        for i in W:
            for t in T:
                cp = _case_pack(problem, i, t)

                model.addConstr(
                    Q_OA[i, t] == cp * q_OA[i, t] + r_OA[i, t],
                    name=f"oa_case_{i}_{t}",
                )

                if cp == 1:
                    model.addConstr(r_OA[i, t] == 0, name=f"oa_res_zero_{i}_{t}")
                    model.addConstr(y_OA[i, t] == 0, name=f"oa_y_zero_{i}_{t}")
                else:
                    model.addConstr(r_OA[i, t] <= (cp - 1), name=f"oa_res_bound_{i}_{t}")
                    model.addConstr(
                        r_OA[i, t] <= (cp - 1) * y_OA[i, t],
                        name=f"oa_res_link_ub_{i}_{t}",
                    )
                    model.addConstr(r_OA[i, t] >= y_OA[i, t], name=f"oa_res_link_lb_{i}_{t}")

        # ====================================================
        # PLT case-pack, activation and transport constraints
        # ====================================================

        for (i, j, t) in arcs:
            cp = _case_pack(problem, j, t)

            model.addConstr(
                Q_PLT[i, j, t] == cp * q_PLT[i, j, t] + r_PLT[i, j, t],
                name=f"plt_case_{i}_{j}_{t}",
            )

            if cp == 1:
                model.addConstr(r_PLT[i, j, t] == 0, name=f"plt_res_zero_{i}_{j}_{t}")
                model.addConstr(y_PLT[i, j, t] == 0, name=f"plt_y_zero_{i}_{j}_{t}")
            else:
                model.addConstr(r_PLT[i, j, t] <= (cp - 1), name=f"plt_res_bound_{i}_{j}_{t}")
                model.addConstr(
                    r_PLT[i, j, t] <= (cp - 1) * y_PLT[i, j, t],
                    name=f"plt_res_link_ub_{i}_{j}_{t}",
                )
                model.addConstr(
                    r_PLT[i, j, t] >= y_PLT[i, j, t],
                    name=f"plt_res_link_lb_{i}_{j}_{t}",
                )

            model.addConstr(
                Q_PLT[i, j, t] <= M * z_PLT[i, j, t],
                name=f"plt_active_ub_{i}_{j}_{t}",
            )

            model.addConstr(
                Q_PLT[i, j, t] >= z_PLT[i, j, t],
                name=f"plt_active_lb_{i}_{j}_{t}",
            )

        # ====================================================
        # Inventory balance
        # ====================================================

        first_t = T[0]

        def oa_arrival_expr(i: str, t: int):
            order_t = _oa_order_period(problem, i, t, arrival_indexing)

            if order_t in T:
                return Q_OA[i, order_t]

            return 0.0

        def plt_in_expr(i: str, t: int):
            terms = []

            for src, dst, ship_t in arcs:
                if dst != i:
                    continue

                lt = int(problem["LT_PLT"][(src, dst)])

                if ship_t + lt == t:
                    terms.append(Q_PLT[src, dst, ship_t])

            return gp.quicksum(terms) if terms else 0.0

        def plt_out_expr(i: str, t: int):
            terms = [
                Q_PLT[src, dst, tt]
                for src, dst, tt in arcs
                if src == i and tt == t
            ]

            return gp.quicksum(terms) if terms else 0.0

        for i in W:
            for idx, t in enumerate(T):
                prev_inventory = problem["BI"].get(i, 0.0) if t == first_t else I[i, T[idx - 1]]

                model.addConstr(
                    I_pre[i, t]
                    == prev_inventory
                    + oa_arrival_expr(i, t)
                    + plt_in_expr(i, t)
                    + problem["delta_I"].get((i, t), 0.0),
                    name=f"pre_inventory_{i}_{t}",
                )

                model.addConstr(
                    I[i, t] == I_pre[i, t] - plt_out_expr(i, t),
                    name=f"end_inventory_{i}_{t}",
                )

        # ====================================================
        # PLT outbound availability: out <= [I_pre - L]^+
        # ====================================================

        for i in W:
            for t in T:
                expr = I_pre[i, t] - problem["L"].get((i, t), 0.0)
                out_expr = plt_out_expr(i, t)

                model.addConstr(available[i, t] >= expr, name=f"avail_lb_expr_{i}_{t}")
                model.addConstr(available[i, t] >= 0, name=f"avail_lb_zero_{i}_{t}")

                model.addConstr(
                    available[i, t] <= expr + M * (1 - available_flag[i, t]),
                    name=f"avail_ub_expr_{i}_{t}",
                )

                model.addConstr(
                    available[i, t] <= M * available_flag[i, t],
                    name=f"avail_ub_zero_{i}_{t}",
                )

                model.addConstr(
                    out_expr <= available[i, t],
                    name=f"plt_out_le_available_{i}_{t}",
                )

        # ====================================================
        # Cost linearization for positive-part terms
        # ====================================================

        for i in W:
            for t in T:
                floor = problem["L"].get((i, t), 0.0)
                ceiling = problem["U"].get((i, t), 0.0)

                model.addConstr(overstock[i, t] >= I[i, t] - ceiling, name=f"overstock_pos_{i}_{t}")
                model.addConstr(shortage[i, t] >= floor - I[i, t], name=f"shortage_pos_{i}_{t}")
                model.addConstr(backorder[i, t] >= -I[i, t], name=f"backorder_pos_{i}_{t}")

        # ====================================================
        # Objective
        # ====================================================

        inventory_cost = gp.quicksum(
            problem["Co"].get((i, t), 0.0) * overstock[i, t]
            + problem["Cs"].get((i, t), 0.0) * shortage[i, t]
            + problem["Cb"].get((i, t), 0.0) * backorder[i, t]
            for i in W
            for t in T
        )

        oa_penalty_cost = gp.quicksum(
            problem["Cp"].get((i, t), 0.0) * y_OA[i, t]
            for i in W
            for t in T
        )

        plt_penalty_cost = gp.quicksum(
            problem["Cp_plt"].get((i, j, t), 0.0) * y_PLT[i, j, t]
            for i, j, t in arcs
        )

        transport_cost = gp.quicksum(
            problem["dist"].get((i, j), 0.0) * problem["TC"] * z_PLT[i, j, t]
            for i, j, t in arcs
        )

        model.setObjective(
            inventory_cost + oa_penalty_cost + plt_penalty_cost + transport_cost,
            GRB.MINIMIZE,
        )

        model.optimize()
        elapsed = time.perf_counter() - start

    except gp.GurobiError as e:
        elapsed = time.perf_counter() - start
        error_message = str(e).lower()
        solver_status = (
            "out_of_memory"
            if "out of memory" in error_message
            else f"gurobi_error_{getattr(e, 'errno', 'unknown')}"
        )

        return _failed_result(product, solver_status, elapsed, str(e), model)

    except Exception as e:
        elapsed = time.perf_counter() - start

        return _failed_result(product, "python_error", elapsed, str(e), model)

    status = _status_name(model)

    if model.SolCount == 0:
        return _failed_result(product, status, elapsed, None, model)

    def val(x) -> float:
        return float(x.X)

    schedule_rows = []
    cost_rows = []
    plt_rows = []

    total_overstock_cost = 0.0
    total_shortage_cost = 0.0
    total_backorder_cost = 0.0
    total_oa_penalty_cost = 0.0
    total_plt_penalty_cost = 0.0
    total_transport_cost = 0.0

    for i in W:
        for t in T:
            inv = val(I[i, t])
            o_qty = val(overstock[i, t])
            s_qty = val(shortage[i, t])
            b_qty = val(backorder[i, t])

            o_cost = problem["Co"].get((i, t), 0.0) * o_qty
            s_cost = problem["Cs"].get((i, t), 0.0) * s_qty
            b_cost = problem["Cb"].get((i, t), 0.0) * b_qty
            oa_pen = problem["Cp"].get((i, t), 0.0) * round(val(y_OA[i, t]))

            total_overstock_cost += o_cost
            total_shortage_cost += s_cost
            total_backorder_cost += b_cost
            total_oa_penalty_cost += oa_pen

            schedule_rows.append(
                {
                    "product": product,
                    "week": t,
                    "warehouse": i,
                    "Q_OA_allocated": round(val(Q_OA[i, t]), 4),
                    "q_OA_case": round(val(q_OA[i, t]), 4),
                    "r_OA_residual": round(val(r_OA[i, t]), 4),
                    "y_OA_penalty": round(val(y_OA[i, t])),
                    "inventory_pre_PLT": round(val(I_pre[i, t]), 4),
                    "inventory_end": round(inv, 4),
                    "overstock_qty": round(o_qty, 4),
                    "shortage_qty": round(s_qty, 4),
                    "backorder_qty": round(b_qty, 4),
                }
            )

            cost_rows.append(
                {
                    "product": product,
                    "week": t,
                    "warehouse": i,
                    "overstock_cost": round(o_cost, 4),
                    "shortage_cost": round(s_cost, 4),
                    "backorder_cost": round(b_cost, 4),
                    "penalty_OA_cost": round(oa_pen, 4),
                    "total_inventory_related_cost": round(o_cost + s_cost + b_cost + oa_pen, 4),
                }
            )

    for i, j, t in arcs:
        qty = val(Q_PLT[i, j, t])

        if qty <= 1e-6:
            continue

        plt_pen = problem["Cp_plt"].get((i, j, t), 0.0) * round(val(y_PLT[i, j, t]))
        trans = problem["dist"].get((i, j), 0.0) * problem["TC"] * round(val(z_PLT[i, j, t]))

        total_plt_penalty_cost += plt_pen
        total_transport_cost += trans

        plt_rows.append(
            {
                "product": product,
                "from_warehouse": i,
                "to_warehouse": j,
                "ship_week": t,
                "lead_time": problem["LT_PLT"].get((i, j), None),
                "arrival_week": t + problem["LT_PLT"].get((i, j), 0),
                "Q_PLT": round(qty, 4),
                "q_PLT_case": round(val(q_PLT[i, j, t]), 4),
                "r_PLT_residual": round(val(r_PLT[i, j, t]), 4),
                "y_PLT_penalty": round(val(y_PLT[i, j, t])),
                "z_PLT_used": round(val(z_PLT[i, j, t])),
                "plt_penalty_cost": round(plt_pen, 4),
                "transport_cost": round(trans, 4),
            }
        )

    total_cost = (
        total_overstock_cost
        + total_shortage_cost
        + total_backorder_cost
        + total_oa_penalty_cost
        + total_plt_penalty_cost
        + total_transport_cost
    )

    best_bound = float(model.ObjBound)
    mip_gap_value = float(model.MIPGap)

    summary = {
        "product": product,
        "solver_status": status,
        "objective_value": round(float(model.ObjVal), 4),
        "recomputed_total_cost": round(total_cost, 4),
        "best_bound": round(best_bound, 4),
        "mip_gap": round(mip_gap_value, 8),
        "runtime_seconds": round(elapsed, 4),
        "num_variables": model.NumVars,
        "num_binary_variables": model.NumBinVars,
        "num_integer_variables": model.NumIntVars,
        "num_constraints": model.NumConstrs,
        "node_count": round(model.NodeCount, 0),
        "simplex_iterations": round(model.IterCount, 0),
        "error_message": None,
        "total_overstock_cost": round(total_overstock_cost, 4),
        "total_shortage_cost": round(total_shortage_cost, 4),
        "total_backorder_cost": round(total_backorder_cost, 4),
        "total_penalty_OA_cost": round(total_oa_penalty_cost, 4),
        "total_plt_penalty_cost": round(total_plt_penalty_cost, 4),
        "total_transport_cost": round(total_transport_cost, 4),
        "grand_total_cost": round(total_cost, 4),
    }

    detail_sheet_rows = build_detail_sheet_rows_from_solution(
        problem=problem,
        T=T,
        W=W,
        arcs=arcs,
        Q_OA=Q_OA,
        q_OA=q_OA,
        r_OA=r_OA,
        Q_PLT=Q_PLT,
        q_PLT=q_PLT,
        r_PLT=r_PLT,
        I_pre=I_pre,
        I=I,
        summary=summary,
        val=val,
        arrival_indexing=arrival_indexing,
    )

    return {
        "product": product,
        "status": status,
        "objective_value": float(model.ObjVal),
        "best_bound": best_bound,
        "mip_gap": mip_gap_value,
        "runtime_seconds": elapsed,
        "schedule_rows": schedule_rows,
        "cost_rows": cost_rows,
        "plt_rows": plt_rows,
        "detail_sheet_rows": detail_sheet_rows,
        "summary": summary,
    }


# ============================================================
# CSV writer
# ============================================================

def write_csv(path: Path, rows: List[Dict], mode: str = "w") -> None:
    if not rows:
        return

    write_header = mode == "w" or not path.exists()

    with path.open(mode, newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))

        if write_header:
            writer.writeheader()

        writer.writerows(rows)


# ============================================================
# Main
# ============================================================

def main() -> None:
    total_start_time = time.perf_counter()

    global DATA_DIR, OUTPUT_DIR

    parser = argparse.ArgumentParser(
        description=(
            "Run exact MILP benchmark for the new SMI-OA-PLT mathematical model. "
            "Capacity mode is fixed to eq."
        )
    )

    parser.add_argument(
        "--data-dir",
        type=str,
        default=str(BASE_DIR / "sample_100_fixed"),
        help="Path to data directory or zip file.",
    )

    parser.add_argument(
        "--output-dir",
        type=str,
        default=str(BASE_DIR / "output_milp_new_model"),
        help="Output directory.",
    )

    parser.add_argument(
        "--max-products",
        type=int,
        default=100,
        help="Number of products to run.",
    )

    parser.add_argument(
        "--time-limit",
        type=float,
        default=None,
        help="Time limit per product in seconds. Omit for no time limit.",
    )

    parser.add_argument(
        "--mip-gap",
        type=float,
        default=0.0,
        help="Target MIP gap. Use 0 for exact optimality.",
    )

    parser.add_argument(
        "--threads",
        type=int,
        default=0,
        help="Number of Gurobi threads. 0 means 2 in this script.",
    )

    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Show Gurobi log.",
    )

    parser.add_argument(
        "--lt-oa",
        type=int,
        default=8,
        help="OA lead time used for all warehouses unless data has a separate source.",
    )

    parser.add_argument(
        "--tc",
        type=float,
        default=1.2,
        help="Transport cost multiplier for PLT route usage.",
    )

    parser.add_argument(
        "--arrival-indexing",
        choices=["business", "math"],
        default="business",
        help=(
            "business: order at period 1 with LT=8 arrives at period 8. "
            "math: uses Q_{t-LT} exactly."
        ),
    )

    args = parser.parse_args()

    if gp is None:
        log.error("gurobipy is not installed. Run: python -m pip install gurobipy")
        sys.exit(1)

    if Workbook is None:
        log.error("openpyxl is not installed. Run: python -m pip install openpyxl")
        sys.exit(1)

    data_path, tmp_dir = _maybe_extract_zip(Path(args.data_dir))

    DATA_DIR = data_path
    OUTPUT_DIR = Path(args.output_dir)

    if not DATA_DIR.exists():
        log.error("Data directory not found: %s", DATA_DIR)
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    schedule_path = OUTPUT_DIR / "milp_schedule.csv"
    cost_path = OUTPUT_DIR / "milp_costs.csv"
    plt_path = OUTPUT_DIR / "milp_plt.csv"
    summary_path = OUTPUT_DIR / "milp_cost_summary.csv"
    run_summary_path = OUTPUT_DIR / "milp_run_summary.csv"
    schedule_detail_path = OUTPUT_DIR / "milp_schedule_detail.xlsx"

    for path in [
        schedule_path,
        cost_path,
        plt_path,
        summary_path,
        run_summary_path,
        schedule_detail_path,
    ]:
        if path.exists():
            path.unlink()

    log.info("Loading data from %s", DATA_DIR)

    product_ids, flow_data = load_inventory_flow(n=args.max_products)
    bi_data = load_inventory_begin()
    costs_data = load_unit_costs()
    cap_data = load_vendor_capacity()
    cp_data = load_packing_details()
    lt_plt = load_plt_lead_times()
    dist_data = load_distances()
    LT_OA = load_oa_lead_times()

    log.info("Found %d products to solve by MILP.", len(product_ids))

    summaries = []
    detail_sheets = {}

    for idx, pid in enumerate(product_ids, 1):
        log.info("=" * 70)
        log.info("[%d/%d] Solving MILP for product %s", idx, len(product_ids), pid)
        log.info("=" * 70)

        pflow = flow_data.get(pid, {})
        bi_map = bi_data.get(pid, {})
        c_map = costs_data.get(pid, {})
        cap_map = cap_data.get(pid, {})
        cp = cp_data.get(pid, 1)

        if not pflow or not bi_map:
            log.warning("Skipping %s because data is incomplete.", pid)
            continue

        problem = build_problem(
            product_id=pid,
            flow=pflow,
            bi_map=bi_map,
            costs_map=c_map,
            cap_map=cap_map,
            cp=cp,
            lt_oa=LT_OA if LT_OA else args.lt_oa,
            lt_plt_map=lt_plt,
            dist_map=dist_data,
            tc=args.tc,
        )

        result = solve_milp_new_model(
            problem=problem,
            time_limit=args.time_limit,
            mip_gap=args.mip_gap,
            threads=args.threads,
            verbose=args.verbose,
            arrival_indexing=args.arrival_indexing,
        )

        summaries.append(result["summary"])

        write_csv(schedule_path, result["schedule_rows"], mode="a")
        write_csv(cost_path, result["cost_rows"], mode="a")
        write_csv(plt_path, result["plt_rows"], mode="a")

        if result.get("detail_sheet_rows"):
            detail_sheets[pid] = result["detail_sheet_rows"]

        log.info("Status   : %s", result["summary"].get("solver_status"))
        log.info("Objective: %s", result["summary"].get("objective_value"))
        log.info("Gap      : %s", result["summary"].get("mip_gap"))
        log.info("Runtime  : %.2f s", result["runtime_seconds"])

    write_csv(summary_path, summaries, mode="w")

    if detail_sheets:
        write_schedule_detail_xlsx(detail_sheets, schedule_detail_path)

    total = sum(
        float(s["grand_total_cost"])
        for s in summaries
        if s.get("grand_total_cost") is not None
    )

    total_runtime_seconds = time.perf_counter() - total_start_time

    solved_count = len([s for s in summaries if s.get("grand_total_cost") is not None])
    optimal_count = len([s for s in summaries if s.get("solver_status") == "optimal"])
    time_limit_count = len([s for s in summaries if s.get("solver_status") == "time_limit"])
    out_of_memory_count = len(
        [s for s in summaries if s.get("solver_status") in {"out_of_memory", "mem_limit"}]
    )
    infeasible_count = len(
        [s for s in summaries if s.get("solver_status") in {"infeasible", "inf_or_unbd", "unbounded"}]
    )
    python_error_count = len([s for s in summaries if s.get("solver_status") == "python_error"])
    failed_count = len(summaries) - solved_count

    run_summary = [
        {
            "n_products_requested": len(product_ids),
            "n_products_solved": solved_count,
            "n_optimal": optimal_count,
            "n_time_limit": time_limit_count,
            "n_out_of_memory": out_of_memory_count,
            "n_infeasible_or_unbounded": infeasible_count,
            "n_python_error": python_error_count,
            "n_failed_or_no_solution": failed_count,
            "total_milp_cost": round(total, 4),
            "total_runtime_seconds": round(total_runtime_seconds, 4),
            "total_runtime_minutes": round(total_runtime_seconds / 60.0, 4),
            "average_runtime_seconds_per_solved_product": round(
                total_runtime_seconds / max(solved_count, 1),
                4,
            ),
            "arrival_indexing": args.arrival_indexing,
            "capacity_mode": "eq",
        }
    ]

    write_csv(run_summary_path, run_summary, mode="w")

    log.info("=" * 70)
    log.info("DONE MILP NEW MODEL")
    log.info("Solved products: %d", solved_count)
    log.info("Optimal products: %d", optimal_count)
    log.info("Time-limit products: %d", time_limit_count)
    log.info("Out-of-memory products: %d", out_of_memory_count)
    log.info("Infeasible/unbounded products: %d", infeasible_count)
    log.info("Python-error products: %d", python_error_count)
    log.info("Failed/no-solution products: %d", failed_count)
    log.info("Total MILP cost: %.4f", total)
    log.info(
        "Total runtime: %.2f seconds = %.2f minutes",
        total_runtime_seconds,
        total_runtime_seconds / 60.0,
    )
    log.info(
        "Average runtime per solved product: %.2f seconds",
        total_runtime_seconds / max(solved_count, 1),
    )
    log.info("Outputs:")

    for path in [
        summary_path,
        schedule_path,
        cost_path,
        plt_path,
        run_summary_path,
        schedule_detail_path,
    ]:
        log.info("  %s", path)

    if tmp_dir is not None:
        tmp_dir.cleanup()


if __name__ == "__main__":
    main()