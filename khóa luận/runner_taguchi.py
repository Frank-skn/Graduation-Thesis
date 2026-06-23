from __future__ import annotations

import argparse
import csv
import json
import logging
import math
import random
import sys
import time
import traceback
import statistics
from collections import defaultdict, Counter
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("taguchi")

# ---------------------------------------------------------------------------
# Import from runner.py
# ---------------------------------------------------------------------------
try:
    from runner import (
        load_inventory_flow, load_inventory_begin, load_unit_costs,
        load_vendor_capacity, load_packing_details, load_plt_lead_times, load_distances,
        build_problem, build_components, extract_results, get_writable_path,
        build_pivot_data, ROW_LABELS, load_oa_lead_times,
    )
except ImportError:
    log.error("Could not import components from runner.py. Make sure runner.py exists in the same folder.")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
DEFAULT_DATA_DIR = BASE_DIR / "sample_100_fixed" 
CONFIG_JSON_PATH = BASE_DIR / "model" / "config.json"
TAGUCHI_CSV_PATH = BASE_DIR / "model" / "taguchi_config.csv"
OUTPUT_DIR = BASE_DIR / "output_taguchi"

# ---------------------------------------------------------------------------
# Excel Styling Helpers
# ---------------------------------------------------------------------------
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def calculate_sn_ratio(values: List[float]) -> float:
    """Calculate Smaller-the-better S/N ratio: -10 * log10( sum(y^2) / n )."""
    if not values:
        return 0.0
    mean_sq = sum(val ** 2 for val in values) / len(values)
    if mean_sq <= 0:
        return 99.99
    return -10.0 * math.log10(mean_sq)


# ---------------------------------------------------------------------------
# Termination Statistics Helpers
# ---------------------------------------------------------------------------
def _get_first_attr(obj: Any, names: List[str], default: Any = None) -> Any:
    """Return the first existing attribute from an object."""
    for name in names:
        if hasattr(obj, name):
            return getattr(obj, name)
    return default


def infer_stop_info(ga: Any, elapsed: float, cfg_dict: Dict[str, Any]) -> Tuple[str, Any, Any]:
    """
    Extract or infer GA termination information.

    Exact mode:
        If GA exposes attributes such as:
            - stop_reason / termination_reason / termination_cause
            - generations_executed / current_generation / generation
            - stagnation_count / stag_count / no_improve_count
        then this function will use them.

    Fallback mode:
        If those attributes do not exist, it infers:
            - TIME_LIMIT if elapsed is close to time_limit_seconds
            - G_MAX if generation counter reaches G_max
            - G_STAG if stagnation counter reaches G_stag
            - UNKNOWN otherwise
    """
    reason = _get_first_attr(
        ga,
        ["stop_reason", "termination_reason", "termination_cause"],
        None,
    )

    n_gen = _get_first_attr(
        ga,
        ["generations_executed", "current_generation", "generation", "n_generations"],
        None,
    )

    stag_count = _get_first_attr(
        ga,
        ["stagnation_count", "stag_count", "no_improve_count"],
        None,
    )

    if isinstance(reason, str) and reason.strip():
        return reason.upper(), n_gen, stag_count

    time_limit = cfg_dict.get("stopping", {}).get("time_limit_seconds", None)
    G_max = cfg_dict.get("ga", {}).get("G_max", None)
    G_stag = cfg_dict.get("ga", {}).get("G_stag", None)

    if time_limit is not None and elapsed >= 0.98 * float(time_limit):
        reason = "TIME_LIMIT"
    elif n_gen is not None and G_max is not None and int(n_gen) >= int(G_max):
        reason = "G_MAX"
    elif stag_count is not None and G_stag is not None and int(stag_count) >= int(G_stag):
        reason = "G_STAG"
    else:
        reason = "UNKNOWN"

    return reason, n_gen, stag_count


def safe_mean(values: List[float], default: float = 0.0) -> float:
    return statistics.mean(values) if values else default


# ---------------------------------------------------------------------------
# Main Taguchi Pipeline
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Taguchi Parameter Tuning Runner for Hybrid GA-ALNS")
    parser.add_argument("--data-dir", type=str, default=str(DEFAULT_DATA_DIR), help="Path to data directory")
    parser.add_argument("--taguchi-csv", type=str, default=str(TAGUCHI_CSV_PATH), help="Path to Taguchi CSV configs")
    parser.add_argument("--max-products", type=int, default=999999, help="Max products to optimize per run")
    parser.add_argument("--repetitions", type=int, default=10, help="Number of repetitions per config")
    parser.add_argument("--no-excel", action="store_true", help="Skip Excel output")
    args, unknown = parser.parse_known_args()

    if unknown:
        log.warning("Ignoring unrecognized arguments: %s", unknown)

    data_dir_path = Path(args.data_dir)
    taguchi_csv_path = Path(args.taguchi_csv)

    if not data_dir_path.exists():
        log.error("Data directory not found: %s", data_dir_path)
        sys.exit(1)

    if not taguchi_csv_path.exists():
        alt_path = BASE_DIR / "taguchi_config.csv"
        if alt_path.exists():
            taguchi_csv_path = alt_path
        else:
            log.error("Taguchi CSV not found: searched %s and %s", taguchi_csv_path, alt_path)
            sys.exit(1)

    if not CONFIG_JSON_PATH.exists():
        log.error("Config JSON template not found: %s", CONFIG_JSON_PATH)
        sys.exit(1)

    # 1. Back up original config.json
    log.info("Backing up model/config.json...")
    with CONFIG_JSON_PATH.open("r", encoding="utf-8-sig") as f:
        original_config_content = f.read()

    # 2. Load Taguchi configs
    taguchi_configs: List[Dict[str, Any]] = []
    with taguchi_csv_path.open("r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            taguchi_configs.append({
                "run_id": int(row["run_id"]),
                "n_pop": int(row["n_pop"]),
                "G_max": int(row["G_max"]),
                "p_crossover": float(row["p_crossover"]),
                "p_mutation": float(row["p_mutation"]),
                "n_iterations": int(row["n_iterations"]),
            })

    log.info("Loaded %d configurations from Taguchi CSV.", len(taguchi_configs))

    # 3. Redirect runner.DATA_DIR
    import runner
    runner.DATA_DIR = data_dir_path

    log.info("Loading problem CSV tables...")
    product_ids, flow_data = load_inventory_flow(n=args.max_products)
    bi_data = load_inventory_begin()
    costs_data = load_unit_costs()
    cap_data = load_vendor_capacity()
    cp_data = load_packing_details()
    lt_plt = load_plt_lead_times()
    dist_data = load_distances()

    LT_OA = load_oa_lead_times()
    TC = 1.2

    log.info(
        "Starting Taguchi Experiment on %d products, %d repetitions per config...",
        len(product_ids),
        args.repetitions,
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    all_runs_data: List[Dict[str, Any]] = []

    try:
        for cfg_idx, tcfg in enumerate(taguchi_configs, 1):
            log.info("=" * 70)
            log.info(
                "[%d/%d] Config ID %d: n_pop=%d, G_max=%d, p_cross=%.2f, p_mut=%.2f, n_iter=%d",
                cfg_idx,
                len(taguchi_configs),
                tcfg["run_id"],
                tcfg["n_pop"],
                tcfg["G_max"],
                tcfg["p_crossover"],
                tcfg["p_mutation"],
                tcfg["n_iterations"],
            )
            log.info("=" * 70)

            # Update model/config.json content
            cfg_dict = json.loads(original_config_content)
            cfg_dict["ga"]["n_pop"] = tcfg["n_pop"]
            cfg_dict["ga"]["G_max"] = tcfg["G_max"]
            cfg_dict["ga"]["p_crossover"] = tcfg["p_crossover"]
            cfg_dict["ga"]["p_mutation"] = tcfg["p_mutation"]
            cfg_dict["alns"]["n_iterations"] = tcfg["n_iterations"]

            # Optional: ensure 10s experiment if config has stopping section.
            # Uncomment if you want to force 10 seconds from this script:
            # cfg_dict.setdefault("stopping", {})["time_limit_seconds"] = 10

            # Save updated config back to file
            with CONFIG_JSON_PATH.open("w", encoding="utf-8") as f:
                json.dump(cfg_dict, f, indent=2)

            for rep in range(1, args.repetitions + 1):
                log.info("  -> Repetition %d/%d (Seed = %d)", rep, args.repetitions, rep)

                rep_costs: List[float] = []
                rep_runtimes: List[float] = []
                rep_generations: List[int] = []
                rep_stag_counts: List[int] = []
                rep_stop_reasons: Counter = Counter()

                success_count = 0

                for idx, pid in enumerate(product_ids, 1):
                    pflow = flow_data.get(pid, {})
                    bi_map = bi_data.get(pid, {})
                    c_map = costs_data.get(pid, {})
                    cap_map = cap_data.get(pid, {})
                    cp = cp_data.get(pid, 1)

                    if not pflow or not bi_map:
                        continue

                    problem = build_problem(
                        product_id=pid,
                        flow=pflow,
                        bi_map=bi_map,
                        costs_map=c_map,
                        cap_map=cap_map,
                        cp=cp,
                        lt_oa=LT_OA,
                        lt_plt_map=lt_plt,
                        dist_map=dist_data,
                        tc=TC,
                    )

                    # Initialize random generator using repetition seed
                    rng = random.Random(rep + idx)
                    ga, _ = build_components(problem, cfg_dict, rng)

                    t0 = time.perf_counter()
                    _, best_sol = ga.run()
                    elapsed = time.perf_counter() - t0

                    stop_reason, n_gen, stag_count = infer_stop_info(ga, elapsed, cfg_dict)

                    rep_costs.append(float(best_sol.fitness))
                    rep_runtimes.append(float(elapsed))
                    rep_stop_reasons[stop_reason] += 1

                    if n_gen is not None:
                        try:
                            rep_generations.append(int(n_gen))
                        except (TypeError, ValueError):
                            pass

                    if stag_count is not None:
                        try:
                            rep_stag_counts.append(int(stag_count))
                        except (TypeError, ValueError):
                            pass

                    success_count += 1

                # Aggregate results across products for this repetition
                if rep_costs:
                    total_cost = sum(rep_costs)
                    avg_cost = total_cost / len(rep_costs)
                    avg_time = sum(rep_runtimes) / len(rep_runtimes)
                    max_time = max(rep_runtimes)

                    dominant_stop_reason = (
                        rep_stop_reasons.most_common(1)[0][0]
                        if rep_stop_reasons
                        else "UNKNOWN"
                    )

                    avg_generations = (
                        statistics.mean(rep_generations)
                        if rep_generations
                        else None
                    )

                    avg_stag_count = (
                        statistics.mean(rep_stag_counts)
                        if rep_stag_counts
                        else None
                    )

                    log.info(
                        "     Total Cost: %12.2f  |  Avg Time: %.3fs  |  Stop: %s",
                        total_cost,
                        avg_time,
                        dominant_stop_reason,
                    )

                    all_runs_data.append({
                        "run_id": tcfg["run_id"],
                        "n_pop": tcfg["n_pop"],
                        "G_max": tcfg["G_max"],
                        "p_crossover": tcfg["p_crossover"],
                        "p_mutation": tcfg["p_mutation"],
                        "n_iterations": tcfg["n_iterations"],
                        "repetition": rep,
                        "seed": rep,
                        "total_cost": round(total_cost, 2),
                        "avg_cost": round(avg_cost, 2),
                        "avg_runtime": round(avg_time, 4),
                        "max_runtime": round(max_time, 4),
                        "success_count": success_count,

                        # Termination statistics per repetition
                        "stop_time_limit": rep_stop_reasons.get("TIME_LIMIT", 0),
                        "stop_gmax": rep_stop_reasons.get("G_MAX", 0),
                        "stop_gstag": rep_stop_reasons.get("G_STAG", 0),
                        "stop_unknown": rep_stop_reasons.get("UNKNOWN", 0),
                        "dominant_stop_reason": dominant_stop_reason,
                        "avg_generations": round(avg_generations, 2) if avg_generations is not None else None,
                        "avg_stag_count": round(avg_stag_count, 2) if avg_stag_count is not None else None,
                    })

    except Exception as e:
        log.error("Error during Taguchi experiment: %s", e)
        log.error(traceback.format_exc())

    finally:
        # Restore original config.json
        log.info("Restoring original model/config.json...")
        with CONFIG_JSON_PATH.open("w", encoding="utf-8") as f:
            f.write(original_config_content)

    # ---------------------------------------------------------------------------
    # Write flat CSV summary of all runs
    # ---------------------------------------------------------------------------
    summary_csv_path = get_writable_path(OUTPUT_DIR / "taguchi_runs.csv")
    if all_runs_data:
        with summary_csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=all_runs_data[0].keys())
            w.writeheader()
            w.writerows(all_runs_data)
        log.info("Saved raw Taguchi runs to CSV: %s", summary_csv_path)

    # ---------------------------------------------------------------------------
    # Calculate Taguchi Aggregated Metrics
    # ---------------------------------------------------------------------------
    log.info("Calculating Taguchi summary metrics...")

    grouped = defaultdict(list)
    for row in all_runs_data:
        grouped[row["run_id"]].append(row)

    taguchi_summaries: List[Dict[str, Any]] = []

    for run_id in sorted(grouped.keys()):
        rows = grouped[run_id]
        costs = [r["total_cost"] for r in rows]
        runtimes = [r["avg_runtime"] for r in rows]

        cfg = rows[0]

        mean_cost = statistics.mean(costs) if costs else 0.0
        std_cost = statistics.stdev(costs) if len(costs) > 1 else 0.0
        min_cost = min(costs) if costs else 0.0
        max_cost = max(costs) if costs else 0.0
        mean_time = statistics.mean(runtimes) if runtimes else 0.0
        sn_ratio = calculate_sn_ratio(costs)

        stop_time_limit = sum(r.get("stop_time_limit", 0) for r in rows)
        stop_gmax = sum(r.get("stop_gmax", 0) for r in rows)
        stop_gstag = sum(r.get("stop_gstag", 0) for r in rows)
        stop_unknown = sum(r.get("stop_unknown", 0) for r in rows)

        stop_total = stop_time_limit + stop_gmax + stop_gstag + stop_unknown

        stop_counter = {
            "TIME_LIMIT": stop_time_limit,
            "G_MAX": stop_gmax,
            "G_STAG": stop_gstag,
            "UNKNOWN": stop_unknown,
        }

        dominant_stop_reason = (
            max(stop_counter, key=stop_counter.get)
            if stop_total > 0
            else "UNKNOWN"
        )

        avg_generations_values = [
            r["avg_generations"]
            for r in rows
            if r.get("avg_generations") is not None
        ]

        avg_stag_values = [
            r["avg_stag_count"]
            for r in rows
            if r.get("avg_stag_count") is not None
        ]

        mean_generations = (
            statistics.mean(avg_generations_values)
            if avg_generations_values
            else 0.0
        )

        mean_stag_count = (
            statistics.mean(avg_stag_values)
            if avg_stag_values
            else 0.0
        )

        taguchi_summaries.append({
            "run_id": run_id,
            "n_pop": cfg["n_pop"],
            "G_max": cfg["G_max"],
            "p_crossover": cfg["p_crossover"],
            "p_mutation": cfg["p_mutation"],
            "n_iterations": cfg["n_iterations"],
            "mean_cost": round(mean_cost, 2),
            "std_cost": round(std_cost, 2),
            "min_cost": round(min_cost, 2),
            "max_cost": round(max_cost, 2),
            "mean_runtime": round(mean_time, 4),
            "sn_ratio": round(sn_ratio, 4),

            # Aggregated termination statistics
            "stop_time_limit": stop_time_limit,
            "stop_gmax": stop_gmax,
            "stop_gstag": stop_gstag,
            "stop_unknown": stop_unknown,
            "pct_time_limit": round(100.0 * stop_time_limit / stop_total, 2) if stop_total else 0.0,
            "pct_gmax": round(100.0 * stop_gmax / stop_total, 2) if stop_total else 0.0,
            "pct_gstag": round(100.0 * stop_gstag / stop_total, 2) if stop_total else 0.0,
            "pct_unknown": round(100.0 * stop_unknown / stop_total, 2) if stop_total else 0.0,
            "dominant_stop_reason": dominant_stop_reason,
            "mean_generations": round(mean_generations, 2),
            "mean_stag_count": round(mean_stag_count, 2),
        })

    # Save summary metrics to CSV
    metrics_csv_path = get_writable_path(OUTPUT_DIR / "taguchi_summary.csv")
    if taguchi_summaries:
        with metrics_csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=taguchi_summaries[0].keys())
            w.writeheader()
            w.writerows(taguchi_summaries)
        log.info("Saved Taguchi summary to CSV: %s", metrics_csv_path)

    # ---------------------------------------------------------------------------
    # Write Excel dashboard report
    # ---------------------------------------------------------------------------
    write_excel = HAS_OPENPYXL and not args.no_excel

    if write_excel and all_runs_data:
        excel_path = get_writable_path(OUTPUT_DIR / "taguchi_analysis.xlsx")
        wb = openpyxl.Workbook()

        # Excel Styles
        CLR_HEADER = "2F5496"
        CLR_SUB = "4472C4"
        CLR_ZEBRA = "F2F5F9"
        CLR_WHITE = "FFFFFF"

        thin = Side(style="thin")
        thick = Side(style="medium")
        grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # --------------------------------------------------------
        # Sheet 1: Taguchi Summary
        # --------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Taguchi Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        headers_summary = [
            "Run ID", "n_pop", "G_max", "p_crossover", "p_mutation", "n_iterations",
            "Mean Total Cost", "StdDev Cost", "Min Cost", "Max Cost", "Mean Runtime (s)",
            "S/N Ratio",
            "Stop TIME_LIMIT", "Stop G_MAX", "Stop G_STAG", "Stop UNKNOWN",
            "% TIME_LIMIT", "% G_MAX", "% G_STAG", "% UNKNOWN",
            "Dominant Stop", "Mean Generations", "Mean Stag Count",
        ]

        for col_idx in range(1, len(headers_summary) + 1):
            ws_summary.column_dimensions[get_column_letter(col_idx)].width = 16

        ws_summary.column_dimensions["A"].width = 8
        ws_summary.column_dimensions["D"].width = 14
        ws_summary.column_dimensions["E"].width = 14
        ws_summary.column_dimensions["F"].width = 14
        ws_summary.column_dimensions["G"].width = 18
        ws_summary.column_dimensions["K"].width = 18
        ws_summary.column_dimensions["U"].width = 18
        ws_summary.column_dimensions["V"].width = 18
        ws_summary.column_dimensions["W"].width = 18

        last_summary_col = get_column_letter(len(headers_summary))

        ws_summary.merge_cells(f"A1:{last_summary_col}1")
        title_cell = ws_summary["A1"]
        title_cell.value = "Taguchi Parameter Design Optimization Analysis"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = _fill(CLR_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[1].height = 40

        ws_summary.merge_cells(f"A2:{last_summary_col}2")
        sub_cell = ws_summary["A2"]
        sub_cell.value = (
            f"Data Source: sample_100_fixed/sample_100_plt_rich "
            f"(Evaluated on first {len(product_ids)} products over {args.repetitions} repetitions)"
        )
        sub_cell.font = Font(size=11, italic=True, color="FFFFFF")
        sub_cell.fill = _fill(CLR_SUB)
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[2].height = 25

        ws_summary.row_dimensions[4].height = 32
        for col_idx, h in enumerate(headers_summary, 1):
            cell = ws_summary.cell(row=4, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _fill(CLR_SUB)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thick, bottom=thick)

        summary_keys = [
            "run_id", "n_pop", "G_max", "p_crossover", "p_mutation", "n_iterations",
            "mean_cost", "std_cost", "min_cost", "max_cost", "mean_runtime",
            "sn_ratio",
            "stop_time_limit", "stop_gmax", "stop_gstag", "stop_unknown",
            "pct_time_limit", "pct_gmax", "pct_gstag", "pct_unknown",
            "dominant_stop_reason", "mean_generations", "mean_stag_count",
        ]

        for row_idx, summary_row in enumerate(taguchi_summaries, 5):
            ws_summary.row_dimensions[row_idx].height = 20
            zebra = CLR_WHITE if row_idx % 2 == 1 else CLR_ZEBRA

            for col_idx, key in enumerate(summary_keys, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=summary_row[key])
                cell.fill = _fill(zebra)
                cell.border = grid_border

                if key in {
                    "run_id", "n_pop", "G_max", "n_iterations",
                    "stop_time_limit", "stop_gmax", "stop_gstag", "stop_unknown",
                }:
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "#,##0"
                elif key in {"p_crossover", "p_mutation"}:
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "0.00"
                elif key in {
                    "mean_cost", "std_cost", "min_cost", "max_cost",
                }:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"
                elif key in {
                    "mean_runtime", "sn_ratio", "pct_time_limit",
                    "pct_gmax", "pct_gstag", "pct_unknown",
                    "mean_generations", "mean_stag_count",
                }:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "0.00"
                else:
                    cell.alignment = Alignment(horizontal="center")

        ws_summary.freeze_panes = ws_summary["A5"]

        # --------------------------------------------------------
        # Sheet 2: All Repetitions Detail
        # --------------------------------------------------------
        ws_details = wb.create_sheet(title="All Repetitions Detail")
        ws_details.views.sheetView[0].showGridLines = True

        headers_detail = [
            "Config ID", "n_pop", "G_max", "p_crossover", "p_mutation", "n_iterations",
            "Repetition", "Seed", "Total Cost", "Avg Cost", "Avg Runtime (s)",
            "Max Runtime (s)", "Total Success",
            "Stop TIME_LIMIT", "Stop G_MAX", "Stop G_STAG", "Stop UNKNOWN",
            "Dominant Stop", "Avg Generations", "Avg Stag Count",
        ]

        for col_idx in range(1, len(headers_detail) + 1):
            ws_details.column_dimensions[get_column_letter(col_idx)].width = 16

        ws_details.column_dimensions["D"].width = 14
        ws_details.column_dimensions["E"].width = 14
        ws_details.column_dimensions["F"].width = 14
        ws_details.column_dimensions["K"].width = 18
        ws_details.column_dimensions["L"].width = 18
        ws_details.column_dimensions["R"].width = 18

        ws_details.row_dimensions[1].height = 32
        for col_idx, h in enumerate(headers_detail, 1):
            cell = ws_details.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = _fill(CLR_SUB)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thick, bottom=thick)

        detail_keys = [
            "run_id", "n_pop", "G_max", "p_crossover", "p_mutation", "n_iterations",
            "repetition", "seed", "total_cost", "avg_cost", "avg_runtime",
            "max_runtime", "success_count",
            "stop_time_limit", "stop_gmax", "stop_gstag", "stop_unknown",
            "dominant_stop_reason", "avg_generations", "avg_stag_count",
        ]

        for row_idx, drow in enumerate(all_runs_data, 2):
            ws_details.row_dimensions[row_idx].height = 18
            zebra = CLR_WHITE if drow["run_id"] % 2 == 1 else CLR_ZEBRA

            for col_idx, key in enumerate(detail_keys, 1):
                cell = ws_details.cell(row=row_idx, column=col_idx, value=drow[key])
                cell.fill = _fill(zebra)
                cell.border = grid_border

                if key in {
                    "run_id", "n_pop", "G_max", "n_iterations",
                    "repetition", "seed", "success_count",
                    "stop_time_limit", "stop_gmax", "stop_gstag", "stop_unknown",
                }:
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "#,##0"
                elif key in {"p_crossover", "p_mutation"}:
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "0.00"
                elif key in {"total_cost", "avg_cost"}:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"
                elif key in {
                    "avg_runtime", "max_runtime", "avg_generations", "avg_stag_count",
                }:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "0.0000"
                else:
                    cell.alignment = Alignment(horizontal="center")

        ws_details.freeze_panes = ws_details["A2"]

        wb.save(str(excel_path))
        log.info("Saved final Taguchi analysis Excel report: %s", excel_path)

    log.info("=" * 70)
    log.info("TAGUCHI PARAMETER SWEEP COMPLETED SUCCESSFULLY.")
    log.info("Taguchi output directory: %s", OUTPUT_DIR)
    log.info("  taguchi_runs.csv     — flat CSV of all runs")
    log.info("  taguchi_summary.csv  — aggregated statistics")
    if write_excel:
        log.info("  taguchi_analysis.xlsx— complete Excel dashboard and analysis")


if __name__ == "__main__":
    main()