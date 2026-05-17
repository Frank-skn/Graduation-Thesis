import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def _val(results: dict, table: str, wh: str, t: int, default=0.0):
    return results.get(table, {}).get(f"{wh}_t{t}", default)

def _plt_val(results: dict, var: str, wi: str, wj: str, t: int):
    # var can be 'Q_PLT', 'q_PLT', 'r_PLT'
    return results.get(var, {}).get(f"{wi}_{wj}_t{t}", 0.0)

def export_master_excel(all_runs: list, output_dir: str):
    """
    Hàm gom toàn bộ chạy batch thành 1 file Master Report.
    all_runs is a list of dicts:
    {
       'case_name': str,
       'solve_time': float,
       'cost_before': float,
       'cost_after': float,
       'results': dict,
       'params': dict
    }
    """
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # SHEET 1: SUMMARY
    # ---------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary Report"
    
    # Summary Metrics
    total_cases = len(all_runs)
    times = [r['solve_time'] for r in all_runs if r['solve_time'] > 0]
    min_time = min(times) if times else 0.0
    max_time = max(times) if times else 0.0
    mean_time = sum(times) / total_cases if total_cases > 0 else 0.0
    total_time = sum(times)
    
    cost_before_total = sum(r['cost_before'] for r in all_runs)
    cost_after_total = sum(r['cost_after'] for r in all_runs)
    
    font_bold = Font(bold=True)
    
    sum_data = [
        ["MASTER EXECUTION SUMMARY"],
        ["Total Cases", total_cases],
        ["Total Solve Time (s)", round(total_time, 2)],
        ["Min Solve Time (s)", round(min_time, 2)],
        ["Max Solve Time (s)", round(max_time, 2)],
        ["Mean Solve Time (s)", round(mean_time, 2)],
        ["Total Net Before Cost", round(cost_before_total, 2)],
        ["Total Net After Cost", round(cost_after_total, 2)],
        []
    ]
    for row in sum_data:
        ws_summary.append(row)
        
    for r in range(1, 9):
        ws_summary.cell(row=r, column=1).font = font_bold
        
    headers = ["Case Name", "Solve Time (s)", "Cost Before (No Optimization)", "Cost After (MILP Optimized)", "Net Savings", "Status"]
    ws_summary.append(headers)
    for c in range(1, len(headers)+1):
        ws_summary.cell(row=10, column=c).font = font_bold
        ws_summary.cell(row=10, column=c).fill = PatternFill(start_color="DCE6F1", fill_type="solid")
    
    for r in all_runs:
        saved = r['cost_before'] - r['cost_after']
        ws_summary.append([
            r['case_name'],
            round(r['solve_time'], 2),
            round(r['cost_before'], 2),
            round(r['cost_after'], 2),
            round(saved, 2),
            r['results'].get('status', 'Unknown')
        ])
        
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 30
    ws_summary.column_dimensions['D'].width = 30
    ws_summary.column_dimensions['E'].width = 20
    ws_summary.column_dimensions['F'].width = 15

    # ---------------------------------------------------------
    # SHEET 2: DETAILED SCHEDULE (Stacked)
    # ---------------------------------------------------------
    ws = wb.create_sheet(title="Schedule Details")
    
    c_red    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    c_green  = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    c_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    c_purple = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
    c_blue   = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    c_orange = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid") # U, L highlight
    
    thick_side = Side(style='thick')
    thin_side  = Side(style='thin')
    top_thick_border = Border(top=thick_side)
    bottom_thick_border = Border(bottom=thick_side)
    
    if total_cases == 0:
        path = os.path.join(output_dir, "master_report.xlsx")
        wb.save(path)
        return

    # Xác định max T để build header (giả sử các case có T bằng nhau hoặc lấy max T)
    max_T = max(run["params"]["meta"]["T"] for run in all_runs)
    
    # Dòng CAP: sẽ ko print gộp ở cột đầu nữa, vì có case name thay thế
    # Heading Header
    headers_sh2 = ["Case Name", "Warehouse", "Data Type", "Case-pack"] + [str(t) for t in range(1, max_T+1)]
    ws.append(headers_sh2)
    for col_idx in range(1, len(headers_sh2) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.border = bottom_thick_border
        cell.font = font_bold
        if col_idx > 4:
            cell.alignment = Alignment(horizontal='center')
            
    current_row = 2
    
    for run in all_runs:
        c_name = run['case_name']
        params = run['params']
        results = run['results']
        
        whs = params["meta"]["warehouses"]
        T   = params["meta"]["T"]
        CP  = params["CP"]
        OAL = params.get("OA_lead", 8)
        BI  = params["BI"]
        
        case_start_row = current_row
        
        # Heading for CAP per case (Optional visual spacing)
        # Bỏ dòng trắng, nhảy thẳng vào
        
        for wh in whs:
            start_row = current_row
            labels = [
                "1. Sàn an toàn (L)",
                "2. Trần tối đa (U)",
                "3. Beginning inventory",
                "4. Inventory fluctuation",
                "5. Net inventory (before)",
                "6. OA Pack qty (arrived)",
                "7. OA Residual (arrived)",
                "8. PLT Pack qty (arrived)",
                "9. PLT Residual (arrived)",
                "10. PLT total sent out",
                "11. Net inventory (after)"
            ]
            
            for idx, lbl in enumerate(labels):
                ws.cell(row=current_row + idx, column=1, value=c_name)
                ws.cell(row=current_row + idx, column=2, value=wh)
                ws.cell(row=current_row + idx, column=3, value=lbl)
                ws.cell(row=current_row + idx, column=4, value=CP)
            
            # Merge cột tên Case, tên Warehouse
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(labels) - 1, end_column=1)
            cname_cell = ws.cell(row=current_row, column=1)
            cname_cell.alignment = Alignment(vertical='center', horizontal='center')
            cname_cell.font = font_bold
            cname_cell.border = top_thick_border
            
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + len(labels) - 1, end_column=2)
            wh_cell = ws.cell(row=current_row, column=2)
            wh_cell.alignment = Alignment(vertical='center', horizontal='center')
            wh_cell.border = top_thick_border
            
            for c in range(1, len(headers_sh2) + 1):
                ws.cell(row=start_row, column=c).border = Border(top=thick_side)
                
            # Điền dữ liệu
            baseline_inv = BI[wh]
            ws.cell(row=current_row + 2, column=5, value=baseline_inv) # Beginning inv t=1 là cột 5 (vì 4 cột đầu label)
            
            for t in range(1, T+1):
                col = 4 + t
                
                L_val = params["L"].get(f"{wh}_{t}", 0)
                U_val = params["U"].get(f"{wh}_{t}", 0)
                
                # 1. L
                ws.cell(row=current_row + 0, column=col, value=L_val).fill = c_orange
                # 2. U
                ws.cell(row=current_row + 1, column=col, value=U_val).fill = c_orange
                
                # 4. Fluctuation
                dI = params["delta_I"].get(f"{wh}_{t}", 0)
                ws.cell(row=current_row + 3, column=col, value=dI)
                
                # 5. Net before
                baseline_inv = baseline_inv + dI
                net_before = baseline_inv
                cell_nb = ws.cell(row=current_row + 4, column=col, value=net_before)
                if net_before < L_val: cell_nb.fill = c_red
                elif net_before > U_val: cell_nb.fill = c_yellow
                else: cell_nb.fill = c_green
                
                # 6. OA Pack
                t_oa = t - OAL
                q_oa_arr = _val(results, "q_OA", wh, t_oa) if 1 <= t_oa <= T else 0
                r_oa_arr = _val(results, "r_OA", wh, t_oa) if 1 <= t_oa <= T else 0
                
                # 7. OA res
                cell_qoa = ws.cell(row=current_row + 5, column=col, value=q_oa_arr)
                cell_qoa.fill = c_purple
                ws.cell(row=current_row + 6, column=col, value=r_oa_arr)
                
                # 8. PLT pack arriving
                q_plt_arr, r_plt_arr = 0, 0
                for wj in whs:
                    if wj != wh:
                        lt = params["PLT_lead"].get(f"{wj}_{wh}", 2)
                        t_plt = t - lt
                        if 1 <= t_plt <= T:
                            q_plt_arr += _plt_val(results, "q_PLT", wj, wh, t_plt)
                            r_plt_arr += _plt_val(results, "r_PLT", wj, wh, t_plt)
                            
                cell_qplt = ws.cell(row=current_row + 7, column=col, value=q_plt_arr)
                if sum([q_plt_arr, r_plt_arr]) > 0:
                    cell_qplt.fill = c_blue
                    ws.cell(row=current_row + 8, column=col, value=r_plt_arr).fill = c_blue
                else:
                    ws.cell(row=current_row + 8, column=col, value=r_plt_arr)
                    
                # 10. PLT sent out
                plt_sent = sum(_plt_val(results, "Q_PLT", wh, wj, t) for wj in whs if wj != wh)
                ws.cell(row=current_row + 9, column=col, value=plt_sent)
                
                # 11. Net after
                net_after = _val(results, "I", wh, t)
                cell_na = ws.cell(row=current_row + 10, column=col, value=net_after)
                if net_after < L_val: cell_na.fill = c_red
                elif net_after > U_val: cell_na.fill = c_yellow
                else: cell_na.fill = c_green
                
            # Đóng biên kho
            for c in range(1, len(headers_sh2) + 1):
                ws.cell(row=current_row + len(labels) - 1, column=c).border = Border(bottom=thick_side)
                
            current_row += len(labels)

    # Chỉnh Auto width
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    for t in range(1, max_T+1):
        ws.column_dimensions[get_column_letter(4 + t)].width = 7
        
    path = os.path.join(output_dir, "master_report.xlsx")
    wb.save(path)
    print(f"  → Đã xuất {path}")
