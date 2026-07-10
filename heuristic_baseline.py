from __future__ import annotations

import argparse
import csv
import json
import logging
import math
import random
import sys
import time
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level  = logging.INFO,
    format = "%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt= "%H:%M:%S",
)
log = logging.getLogger("heuristic")

# ---------------------------------------------------------------------------
# Internal imports
# ---------------------------------------------------------------------------
from model.core.problem import Problem

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output_baseline"


# ---------------------------------------------------------------------------
# Heuristic Helpers
# ---------------------------------------------------------------------------

def _safe(mapping: Dict, key: Any, default: float = 0.0) -> float:
    """
    Lấy giá trị an toàn từ dictionary.
    Nếu key không tồn tại hoặc lỗi kiểu dữ liệu thì trả về default.
    """
    try:
        return float(mapping.get(key, default))
    except Exception:
        return float(default)


def _case_pack(problem, wh: str, t: int) -> int:
    """
    Lấy case-pack của sản phẩm.

    Trong dữ liệu hiện tại, CP thường là một số chung cho sản phẩm.
    Hàm này vẫn hỗ trợ trường hợp CP là dict theo warehouse/time.
    """
    cp = getattr(problem, "CP", 1)

    if isinstance(cp, dict):
        value = cp.get((wh, t)) or cp.get(wh) or cp.get(t) or 1
    else:
        value = cp

    return max(int(round(float(value))), 1)


def _floor_case(qty: float, cp: int) -> int:
    """
    Làm tròn xuống theo case-pack.
    Ví dụ: qty = 23, cp = 6 thì kết quả = 18.
    """
    return max(0, int(qty // cp) * cp)


def _residual(qty: float, cp: int) -> float:
    """
    Tính phần dư sau khi chia theo case-pack.
    Nếu phần dư > 0 thì có thể phát sinh penalty case-pack.
    """
    return max(0.0, qty - math.floor(qty / cp) * cp)


# ---------------------------------------------------------------------------
# Greedy Operational Heuristic Class
# ---------------------------------------------------------------------------

class GreedyOperationalHeuristic:
    """
    Core Greedy Operational Heuristic.

    Thuật toán mô phỏng cách vận hành thực tế:
    1. Cập nhật tồn kho theo từng giai đoạn.
    2. Điều chuyển ngang PLT từ nhà máy dư sang nhà máy thiếu.
    3. Phân bổ OA từ nguồn cung trung tâm theo mức thiếu hụt dự kiến.
    4. Tính tổng chi phí vận hành.
    """

    def __init__(self, problem):
        self.p = problem
        self.warehouses = list(problem.warehouses)
        self.periods = list(problem.periods)
        self.max_t = max(self.periods)

        self.Q_oa: Dict[Tuple[str, int], float] = {}
        self.Q_plt: Dict[Tuple[str, str, int], float] = {}
        self.inventory: Dict[Tuple[str, int], float] = {}

    def solve(self) -> Dict[str, Any]:
        """
        Hàm chính của thuật toán.

        Với mỗi giai đoạn t:
        - Cập nhật tồn kho tạm thời.
        - Thực hiện PLT nếu có nhà máy dư và nhà máy thiếu.
        - Lưu tồn kho cuối kỳ.
        - Phân bổ OA cho kỳ hiện tại.
        """
        start_time = time.perf_counter()

        current_inventory = {
            wh: _safe(self.p.BI, wh, 0.0)
            for wh in self.warehouses
        }

        for t in self.periods:
            temp_inventory: Dict[str, float] = {}

            # 1. Cập nhật tồn kho trước PLT
            for wh in self.warehouses:
                temp_inventory[wh] = (
                    current_inventory.get(wh, 0.0)
                    + self._arrive_oa(wh, t)
                    + self._arrive_plt(wh, t)
                    + _safe(self.p.delta_I, (wh, t), 0.0)
                )

            # 2. Điều chuyển ngang PLT
            self._greedy_plt(t, temp_inventory)

            # 3. Lưu tồn kho cuối kỳ
            for wh in self.warehouses:
                self.inventory[(wh, t)] = temp_inventory[wh]

            current_inventory = temp_inventory

            # 4. Phân bổ OA cho kỳ t
            oa_plan = self._greedy_oa(t, current_inventory)
            for wh, qty in oa_plan.items():
                self.Q_oa[(wh, t)] = qty

        # 5. Tính objective
        breakdown = self._evaluate()

        return {
            "objective_value": round(breakdown["total_cost"], 4),
            "runtime_seconds": round(time.perf_counter() - start_time, 4),
            "solver_status": "ok",
            "breakdown": breakdown,
            "Q_oa": self.Q_oa,
            "Q_plt": self.Q_plt,
            "inventory": self.inventory,
        }

    # ============================================================
    # Arrival functions
    # ============================================================

    def _arrive_oa(self, wh: str, t: int) -> float:
        """
        Tính lượng OA đến warehouse wh tại kỳ t.

        Nếu OA đặt tại kỳ s thì hàng đến tại:
            s + LT_OA
        """
        lt = int(getattr(self.p, "LT_OA", 0))
        order_t = t - lt

        if order_t < min(self.periods):
            return 0.0

        return float(self.Q_oa.get((wh, order_t), 0.0))

    def _arrive_plt(self, wh: str, t: int) -> float:
        """
        Tính lượng PLT đến warehouse wh tại kỳ t.

        Nếu PLT gửi tại kỳ s từ donor đến receiver thì hàng đến tại:
            s + LT_PLT[donor, receiver]
        """
        total = 0.0

        for (donor, receiver, ship_t), qty in self.Q_plt.items():
            if receiver != wh:
                continue

            lt = int(self.p.LT_PLT.get((donor, receiver), 0))
            if ship_t + lt == t:
                total += qty

        return total

    # ============================================================
    # OA allocation
    # ============================================================

    def _greedy_oa(self, t: int, current_inventory: Dict[str, float]) -> Dict[str, int]:
        """
        Phân bổ OA theo mức thiếu hụt dự kiến.

        Ý tưởng:
        - Dự báo tồn kho tại thời điểm hàng OA đến.
        - Nhà máy nào có nguy cơ thiếu nhiều hơn thì được ưu tiên nhận hàng.
        - Lượng phân bổ được làm tròn theo case-pack.
        """
        capacity = int(round(_safe(self.p.CAP, t, 0.0)))

        if capacity <= 0:
            return {wh: 0 for wh in self.warehouses}

        priorities = self._oa_priorities(t, current_inventory)
        total_priority = sum(priorities.values())

        if total_priority <= 0:
            priorities = {wh: 1.0 for wh in self.warehouses}
            total_priority = float(len(self.warehouses))

        allocation = {
            wh: _floor_case(
                capacity * priorities[wh] / total_priority,
                _case_pack(self.p, wh, t),
            )
            for wh in self.warehouses
        }

        allocation = self._repair_capacity(
            allocation=allocation,
            capacity=capacity,
            priorities=priorities,
            t=t,
        )

        return allocation

    def _oa_priorities(self, t: int, current_inventory: Dict[str, float]) -> Dict[str, float]:
        """
        Tính điểm ưu tiên nhận OA.

        Priority cao nếu:
        - tồn kho dự kiến thấp hơn inventory floor,
        - có nguy cơ backorder.
        """
        lt = int(getattr(self.p, "LT_OA", 0))
        target_t = min(t + lt, self.max_t)

        priorities: Dict[str, float] = {}

        for wh in self.warehouses:
            projected_inventory = self._project_inventory(
                wh=wh,
                current_t=t,
                target_t=target_t,
                current_inventory=current_inventory.get(wh, 0.0),
            )

            floor = _safe(self.p.L, (wh, target_t), 0.0)
            shortage_cost = _safe(self.p.Cs, (wh, target_t), 1.0)
            backorder_cost = _safe(self.p.Cb, (wh, target_t), 1.0)

            shortage_need = max(0.0, floor - projected_inventory)
            backorder_risk = max(0.0, -projected_inventory)

            priorities[wh] = max(
                shortage_need * shortage_cost
                + backorder_risk * backorder_cost,
                0.0,
            )

        return priorities

    def _project_inventory(
        self,
        wh: str,
        current_t: int,
        target_t: int,
        current_inventory: float,
    ) -> float:
        """
        Dự báo tồn kho từ current_t đến target_t.

        Dùng để đánh giá nguy cơ thiếu khi hàng OA hoặc PLT đến.
        """
        projected = float(current_inventory)
        lt_oa = int(getattr(self.p, "LT_OA", 0))

        for tau in range(current_t + 1, target_t + 1):
            projected += _safe(self.p.delta_I, (wh, tau), 0.0)

            # OA đã đặt trước đó và đến tại tau
            order_t = tau - lt_oa
            if order_t in self.periods:
                projected += self.Q_oa.get((wh, order_t), 0.0)

            # PLT đã đặt trước đó và đến tại tau
            for (donor, receiver, ship_t), qty in self.Q_plt.items():
                if receiver != wh:
                    continue

                lt_plt = int(self.p.LT_PLT.get((donor, receiver), 0))
                if ship_t + lt_plt == tau:
                    projected += qty

        return projected

    def _repair_capacity(
        self,
        allocation: Dict[str, int],
        capacity: int,
        priorities: Dict[str, float],
        t: int,
    ) -> Dict[str, int]:
        """
        Sửa tổng OA để đảm bảo tổng phân bổ bằng capacity.

        Sau khi làm tròn theo case-pack, tổng phân bổ có thể nhỏ hơn capacity.
        Phần còn lại được phân bổ cho nhà máy có priority cao.
        """
        remain = capacity - sum(allocation.values())

        if remain <= 0:
            return allocation

        sorted_warehouses = sorted(
            self.warehouses,
            key=lambda wh: priorities.get(wh, 0.0),
            reverse=True,
        )

        changed = True

        while remain > 0 and changed:
            changed = False

            for wh in sorted_warehouses:
                if remain <= 0:
                    break

                cp = _case_pack(self.p, wh, t)

                if remain >= cp:
                    allocation[wh] += cp
                    remain -= cp
                    changed = True

        if remain > 0:
            allocation[sorted_warehouses[0]] += remain

        return allocation

    # ============================================================
    # PLT allocation
    # ============================================================

    def _greedy_plt(self, t: int, inventory_now: Dict[str, float]) -> None:
        """
        Điều chuyển PLT theo logic dư - thiếu nhìn trước.

        Donor:
            warehouse có tồn kho hiện tại cao hơn floor.

        Receiver:
            warehouse có nguy cơ thiếu tại thời điểm hàng PLT đến.
        """
        surplus = {
            wh: max(
                0.0,
                inventory_now.get(wh, 0.0) - _safe(self.p.L, (wh, t), 0.0),
            )
            for wh in self.warehouses
        }

        surplus = {
            wh: qty
            for wh, qty in surplus.items()
            if qty > 0
        }

        if not surplus:
            return

        receiver_scores = self._plt_scores(t, inventory_now)

        receivers = sorted(
            [wh for wh, score in receiver_scores.items() if score > 0],
            key=lambda wh: receiver_scores[wh],
            reverse=True,
        )

        for receiver in receivers:
            if t not in self.p.PLT_periods.get(receiver, frozenset()):
                continue

            donors = sorted(
                surplus.keys(),
                key=lambda donor: (
                    -surplus.get(donor, 0.0),
                    self.p.dist.get((donor, receiver), 1e9),
                ),
            )

            for donor in donors:
                if donor == receiver:
                    continue

                available = surplus.get(donor, 0.0)

                if available <= 0:
                    continue

                if (donor, receiver) not in self.p.LT_PLT:
                    continue

                lt = int(self.p.LT_PLT[(donor, receiver)])
                arrival_t = t + lt

                if arrival_t > self.max_t:
                    continue

                projected_receiver = self._project_inventory(
                    wh=receiver,
                    current_t=t,
                    target_t=arrival_t,
                    current_inventory=inventory_now.get(receiver, 0.0),
                )

                need = max(
                    0.0,
                    _safe(self.p.L, (receiver, arrival_t), 0.0)
                    - projected_receiver,
                )

                if need <= 0:
                    continue

                transfer = _floor_case(
                    min(available, need),
                    _case_pack(self.p, receiver, arrival_t),
                )

                if transfer <= 0:
                    continue

                key = (donor, receiver, t)
                self.Q_plt[key] = self.Q_plt.get(key, 0.0) + transfer

                # Donor mất hàng ngay tại kỳ t
                inventory_now[donor] -= transfer
                surplus[donor] -= transfer

                # Nếu lead time = 0 thì receiver nhận ngay
                if lt == 0:
                    inventory_now[receiver] += transfer

                if surplus[donor] <= 0:
                    break

    def _plt_scores(self, t: int, inventory_now: Dict[str, float]) -> Dict[str, float]:
        """
        Tính điểm ưu tiên nhận PLT.

        Điểm cao nếu warehouse có nguy cơ thiếu khi hàng PLT đến.
        """
        scores: Dict[str, float] = {}

        for receiver in self.warehouses:
            if t not in self.p.PLT_periods.get(receiver, frozenset()):
                scores[receiver] = 0.0
                continue

            best_score = 0.0

            for donor in self.warehouses:
                if donor == receiver:
                    continue

                if (donor, receiver) not in self.p.LT_PLT:
                    continue

                lt = int(self.p.LT_PLT[(donor, receiver)])
                arrival_t = t + lt

                if arrival_t > self.max_t:
                    continue

                projected = self._project_inventory(
                    wh=receiver,
                    current_t=t,
                    target_t=arrival_t,
                    current_inventory=inventory_now.get(receiver, 0.0),
                )

                need = max(
                    0.0,
                    _safe(self.p.L, (receiver, arrival_t), 0.0) - projected,
                )

                shortage_cost = _safe(self.p.Cs, (receiver, arrival_t), 1.0)
                best_score = max(best_score, need * shortage_cost)

            scores[receiver] = best_score

        return scores

    # ============================================================
    # Objective evaluation
    # ============================================================

    def _evaluate(self) -> Dict[str, float]:
        """
        Tính objective của nghiệm.

        Bao gồm:
        - inventory cost: overstock, shortage, backorder;
        - OA penalty nếu không đúng case-pack;
        - PLT penalty nếu không đúng case-pack;
        - transport cost nếu có PLT.
        """
        p = self.p

        inventory_cost = 0.0
        oa_penalty_cost = 0.0
        plt_penalty_cost = 0.0
        transport_cost = 0.0

        for wh in self.warehouses:
            for t in self.periods:
                inv = self.inventory.get((wh, t), 0.0)

                floor = _safe(p.L, (wh, t), 0.0)
                ceiling = _safe(p.U, (wh, t), 0.0)

                overstock = max(0.0, inv - ceiling)
                shortage = max(0.0, floor - inv)
                backorder = max(0.0, -inv)

                inventory_cost += (
                    _safe(p.Co, (wh, t)) * overstock
                    + _safe(p.Cs, (wh, t)) * shortage
                    + _safe(p.Cb, (wh, t)) * backorder
                )

        for (wh, t), qty in self.Q_oa.items():
            if _residual(qty, _case_pack(p, wh, t)) > 1e-9:
                oa_penalty_cost += _safe(p.Cp, (wh, t))

        for (donor, receiver, t), qty in self.Q_plt.items():
            if qty <= 0:
                continue

            if _residual(qty, _case_pack(p, receiver, t)) > 1e-9:
                plt_penalty_cost += _safe(p.Cp_plt, (donor, receiver, t))

            transport_cost += p.dist.get((donor, receiver), 0.0) * float(p.TC)

        total_cost = (
            inventory_cost
            + oa_penalty_cost
            + plt_penalty_cost
            + transport_cost
        )

        return {
            "total_cost": total_cost,
            "inventory_cost": inventory_cost,
            "oa_penalty_cost": oa_penalty_cost,
            "plt_penalty_cost": plt_penalty_cost,
            "transport_cost": transport_cost,
        }


# ===========================================================================
# Heuristic Solution Wrapper to reuse Excel and CSV builders of runner.py
# ===========================================================================

class HeuristicSolutionWrapper:
    def __init__(self, problem: Problem, sol_dict: Dict[str, Any]):
        self.I = sol_dict["inventory"]
        self.Q_OA = sol_dict["Q_oa"]
        self.Q_PLT = sol_dict["Q_plt"]
        self.fitness = sol_dict["objective_value"]
        
        # Build r_OA (1 if there's fractional residual, 0 otherwise)
        self.r_OA = {}
        for wh in problem.warehouses:
            for t in problem.periods:
                qty = self.Q_OA.get((wh, t), 0.0)
                cp = _case_pack(problem, wh, t)
                self.r_OA[(wh, t)] = 1 if _residual(qty, cp) > 1e-9 else 0


# ---------------------------------------------------------------------------
# Data loaders (Matches runner.py with utf-8-sig and _format_wh robustness)
# ---------------------------------------------------------------------------

def load_inventory_flow(n: int = 999999) -> Tuple[List[str], Dict]:
    path = DATA_DIR / "inventory_flow.csv"
    product_order: List[str] = []
    flow: Dict[str, Dict[str, Dict[int, Dict]]] = defaultdict(lambda: defaultdict(dict))

    with path.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = row["product_id"]
            wh  = row["warehouse_id"]
            t   = int(row["time_period"])
            if pid not in product_order:
                if len(product_order) >= n:
                    continue
                product_order.append(pid)
            if pid in product_order:
                flow[pid][wh][t] = {
                    "delta_I": float(row["inventory_fluctuation"]),
                    "U"      : float(row["inventory_ceiling"]),
                    "L"      : float(row["inventory_floor"]),
                }
    return product_order, flow


def load_inventory_begin() -> Dict[str, Dict[str, float]]:
    bi: Dict[str, Dict[str, float]] = defaultdict(dict)
    with (DATA_DIR / "inventory_begin.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            bi[row["product_id"]][row["warehouse_id"]] = float(row["beginning_inventory"])
    return bi


def load_unit_costs() -> Dict[str, Dict[str, Dict[int, Dict]]]:
    costs: Dict = defaultdict(lambda: defaultdict(dict))
    with (DATA_DIR / "unit_cost.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            pid = row["product_id"]
            wh  = row["warehouse_id"]
            t   = int(row["time_period"])
            costs[pid][wh][t] = {
                "Co": float(row["overstock_cost"]),
                "Cs": float(row["shortage_cost"]),
                "Cb": float(row["backlog_cost"]),
                "Cp": float(row["penalty_cost"]),
            }
    return costs


def load_vendor_capacity() -> Dict[str, Dict[int, float]]:
    cap: Dict[str, Dict[int, float]] = defaultdict(dict)
    with (DATA_DIR / "vendor_capacity.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cap[row["product_id"]][int(row["time_period"])] = float(row["capacity"])
    return cap


def load_packing_details() -> Dict[str, int]:
    cp: Dict[str, int] = {}
    with (DATA_DIR / "packing_details.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            cp[row["product_id"]] = int(row["pack_multiple"])
    return cp


def _format_wh(val: str) -> str:
    val = val.strip()
    if val.startswith("WH"):
        return val
    try:
        return f"WH{int(val):02d}"
    except ValueError:
        return val


def load_plt_lead_times() -> Dict[Tuple[str, str], int]:
    lt: Dict[Tuple[str, str], int] = {}
    with (DATA_DIR / "plt_lead_time.csv").open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            frm = _format_wh(row["from_warehouse_id"])
            to  = _format_wh(row["to_warehouse_id"])
            lt[(frm, to)] = int(row["lead_time_weeks"])
    return lt


def load_distances() -> Dict[Tuple[str, str], float]:
    state_map = {
        "MI": "WH01", "OH": "WH02", "IN": "WH03",
        "IL": "WH04", "KY": "WH05", "MO": "WH06",
    }
    dist: Dict[Tuple[str, str], float] = {}
    with (DATA_DIR / "FGPs_distance.csv").open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            from_state = row["State"].strip()
            from_wh = state_map.get(from_state)
            if not from_wh:
                continue
            for state, wh in state_map.items():
                if state in row:
                    dist[(from_wh, wh)] = float(row[state])
    return dist


# ---------------------------------------------------------------------------
# Problem Builder Factory
# ---------------------------------------------------------------------------

def build_problem(
    product_id  : str,
    flow        : Dict[str, Dict[int, Dict]],
    bi_map      : Dict[str, float],
    costs_map   : Dict[str, Dict[int, Dict]],
    cap_map     : Dict[int, float],
    cp          : int,
    lt_oa       : int,
    lt_plt_map  : Dict[Tuple[str, str], int],
    dist_map    : Dict[Tuple[str, str], float],
    tc          : float = 1.2,
) -> Problem:
    warehouses = tuple(sorted(flow.keys()))
    periods_set = set()
    for wh_data in flow.values():
        periods_set.update(wh_data.keys())
    periods = tuple(sorted(periods_set))

    LT_PLT = {(i, j): v for (i, j), v in lt_plt_map.items() if i in warehouses and j in warehouses}

    plt_cutoff = lt_oa - 1
    PLT_periods = {
        wh: frozenset(t for t in periods if t <= plt_cutoff)
        for wh in warehouses
    }

    BI      = {wh: bi_map.get(wh, 0.0) for wh in warehouses}
    delta_I = {}
    U = {}
    L = {}
    for wh in warehouses:
        for t in periods:
            entry = flow[wh].get(t, {})
            delta_I[(wh, t)] = entry.get("delta_I", 0.0)
            U[(wh, t)] = entry.get("U", 1e6)
            L[(wh, t)] = entry.get("L", 0.0)

    CAP = {t: cap_map.get(t, 9999.0) for t in periods}

    Co = {}
    Cs = {}
    Cb = {}
    Cp_map = {}
    for wh in warehouses:
        for t in periods:
            c = costs_map.get(wh, {}).get(t, {})
            Co[(wh, t)] = c.get("Co", 0.1)
            Cs[(wh, t)] = c.get("Cs", 0.5)
            Cb[(wh, t)] = c.get("Cb", 1500.0)
            Cp_map[(wh, t)] = c.get("Cp", 2000.0)

    dist = {(i, j): dist_map.get((i, j), 0.0) for i in warehouses for j in warehouses}

    Cp_plt = {}
    for (i, j) in LT_PLT:
        for t in periods:
            Cp_plt[(i, j, t)] = Cp_map.get((j, t), 0.0)

    return Problem(
        product     = product_id,
        warehouses  = warehouses,
        periods     = periods,
        LT_OA       = lt_oa,
        LT_PLT      = LT_PLT,
        PLT_periods = PLT_periods,
        BI          = BI,
        delta_I     = delta_I,
        U           = U,
        L           = L,
        CAP         = CAP,
        CP          = cp,
        TC          = tc,
        dist        = dist,
        Co          = Co,
        Cs          = Cs,
        Cb          = Cb,
        Cp          = Cp_map,
        Cp_plt      = Cp_plt,
    )


# ---------------------------------------------------------------------------
# Results extractors (Mirroring runner.py perfectly)
# ---------------------------------------------------------------------------

def extract_results(problem: Problem, sol: HeuristicSolutionWrapper, elapsed: float):
    p = problem
    schedule_rows = []
    cost_rows     = []
    grand = {k: 0.0 for k in ["overstock", "shortage", "backorder", "penalty_OA", "transport", "total"]}

    for t in p.periods:
        for wh in p.warehouses:
            inv   = sol.I.get((wh, t), 0.0)
            q_oa  = sol.Q_OA.get((wh, t), 0)
            lt_oa = p.LT_OA
            q_received = sol.Q_OA.get((wh, t - lt_oa), 0) if (t - lt_oa) in p.periods else 0

            plt_in = sum(
                sol.Q_PLT.get((src, wh, t2), 0.0)
                for src in p.warehouses if src != wh
                for t2 in p.periods if t2 == t - p.LT_PLT.get((src, wh), 0)
            )
            plt_out = sum(sol.Q_PLT.get((wh, dst, t), 0.0) for dst in p.warehouses if dst != wh)

            u_val = p.U.get((wh, t), 1e9)
            l_val = p.L.get((wh, t), 0.0)
            Co_   = p.Co.get((wh, t), 0.0) * max(inv - u_val, 0.0)
            Cs_   = p.Cs.get((wh, t), 0.0) * max(l_val - inv, 0.0)
            Cb_   = p.Cb.get((wh, t), 0.0) * max(-inv, 0.0)
            r_val = sol.r_OA.get((wh, t), 0)
            Cp_   = p.Cp.get((wh, t), 0.0) if r_val > 0 else 0.0

            tc_row = sum(
                p.dist.get((i, j), 0.0) * p.TC
                for (i, j, t2), q in sol.Q_PLT.items()
                if t2 == t and j == wh and q > 0
            )

            row_total = Co_ + Cs_ + Cb_ + Cp_ + tc_row

            inv_begin = sol.I.get((wh, t - 1), p.BI.get(wh, 0.0)) if t > p.periods[0] else p.BI.get(wh, 0.0)

            schedule_rows.append({
                "product"        : p.product,
                "week"           : t,
                "warehouse"      : wh,
                "inv_begin"      : round(inv_begin, 2),
                "Q_OA_allocated" : q_oa,
                "Q_OA_received"  : q_received,
                "PLT_in"         : round(plt_in, 2),
                "PLT_out"        : round(plt_out, 2),
                "delta_I"        : p.delta_I.get((wh, t), 0.0),
                "inventory_end"  : round(inv, 2),
                "surplus_E"      : round(max(inv - l_val, 0.0), 2),
                "backorder"      : round(max(-inv, 0.0), 2),
                "shortage"       : round(max(l_val - inv, 0.0), 2),
                "overstock"      : round(max(inv - u_val, 0.0), 2),
            })

            cost_rows.append({
                "product"          : p.product,
                "week"             : t,
                "warehouse"        : wh,
                "overstock_cost"   : round(Co_, 4),
                "shortage_cost"    : round(Cs_, 4),
                "backorder_cost"   : round(Cb_, 4),
                "penalty_OA_cost"  : round(Cp_, 4),
                "transport_cost"   : round(tc_row, 4),
                "total_cost"       : round(row_total, 4),
            })

            grand["overstock"]  += Co_
            grand["shortage"]   += Cs_
            grand["backorder"]  += Cb_
            grand["penalty_OA"] += Cp_
            grand["transport"]  += tc_row
            grand["total"]      += row_total

    summary = {
        "product"                : p.product,
        "n_warehouses"           : p.n_wh,
        "n_periods"              : p.n_periods,
        "fitness"                : round(sol.fitness, 4),
        "elapsed_s"              : round(elapsed, 4),
        "total_overstock_cost"   : round(grand["overstock"],  4),
        "total_shortage_cost"    : round(grand["shortage"],   4),
        "total_backorder_cost"   : round(grand["backorder"],  4),
        "total_penalty_OA_cost"  : round(grand["penalty_OA"], 4),
        "total_transport_cost"   : round(grand["transport"],  4),
        "grand_total_cost"       : round(grand["total"],      4),
    }

    return schedule_rows, cost_rows, summary


# ---------------------------------------------------------------------------
# CSV and Excel Pivot writers (Mirroring runner.py perfectly)
# ---------------------------------------------------------------------------

def write_csv(path: Path, rows: List[Dict], mode: str = "w") -> None:
    if not rows:
        return
    write_header = mode == "w" or not path.exists()
    with path.open(mode, newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        if write_header:
            w.writeheader()
        w.writerows(rows)


def get_writable_path(path: Path) -> Path:
    """If the file is locked by Excel, append _1, _2... until writable."""
    if not path.exists():
        return path
    try:
        with path.open("a", encoding="utf-8-sig") as f:
            pass
        return path
    except (IOError, PermissionError):
        suffix = 1
        while True:
            alt_path = path.parent / f"{path.stem}_{suffix}{path.suffix}"
            if not alt_path.exists():
                return alt_path
            try:
                with alt_path.open("a", encoding="utf-8-sig") as f:
                    pass
                return alt_path
            except (IOError, PermissionError):
                suffix += 1


ROW_LABELS = [
    "Beginning inventory",
    "Inventory fluctuation",
    "Net inventory (before)",
    "OA Pack qty (arrived)",
    "OA Residual (arrived)",
    "PLT Pack qty (arrived)",
    "PLT Residual (arrived)",
    "PLT total sent out",
    "Net inventory (after)",
]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def build_pivot_data(problem: Problem, sol: HeuristicSolutionWrapper) -> Dict:
    p   = problem
    out = {}

    for wh in p.warehouses:
        rows: Dict[str, Dict[int, float]] = {lbl: {} for lbl in ROW_LABELS}

        for t in p.periods:
            inv_begin = (
                sol.I.get((wh, t - 1), p.BI.get(wh, 0.0))
                if t > p.periods[0]
                else p.BI.get(wh, 0.0)
            )
            delta     = p.delta_I.get((wh, t), 0.0)
            net_before = inv_begin + delta

            lt_oa = p.LT_OA
            q_oa_recv = sol.Q_OA.get((wh, t - lt_oa), 0) if (t - lt_oa) in p.periods else 0
            oa_pack = math.floor(q_oa_recv / p.CP) * p.CP if p.CP > 0 else q_oa_recv
            oa_resid = q_oa_recv - oa_pack

            plt_recv_total = sum(
                sol.Q_PLT.get((src, wh, t2), 0.0)
                for src in p.warehouses if src != wh
                for t2 in p.periods
                if t2 == t - p.LT_PLT.get((src, wh), 0)
            )
            plt_pack  = math.floor(plt_recv_total / p.CP) * p.CP if p.CP > 0 else plt_recv_total
            plt_resid = plt_recv_total - plt_pack

            plt_sent = sum(
                sol.Q_PLT.get((wh, dst, t), 0.0)
                for dst in p.warehouses if dst != wh
            )

            inv_after = sol.I.get((wh, t), 0.0)

            rows["Beginning inventory"][t]    = round(inv_begin, 1)
            rows["Inventory fluctuation"][t]  = round(delta, 1)
            rows["Net inventory (before)"][t] = round(net_before, 1)
            rows["OA Pack qty (arrived)"][t]  = round(oa_pack, 1)
            rows["OA Residual (arrived)"][t]  = round(oa_resid, 1)
            rows["PLT Pack qty (arrived)"][t] = round(plt_pack, 1)
            rows["PLT Residual (arrived)"][t] = round(plt_resid, 1)
            rows["PLT total sent out"][t]     = round(plt_sent, 1)
            rows["Net inventory (after)"][t]  = round(inv_after, 1)

        out[wh] = rows
    return out


def add_product_sheet(wb: openpyxl.Workbook, problem: Problem, sol: HeuristicSolutionWrapper, summary: Dict) -> None:
    if not HAS_OPENPYXL:
        return

    p       = problem
    periods = list(p.periods)
    n_per   = len(periods)

    sheet_name = p.product[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Color palette
    CLR_HEADER_DARK  = "2F5496"
    CLR_HEADER_MID   = "4472C4"
    CLR_WH_LABEL     = "D6E4F0"
    CLR_NET_BEFORE   = "FFF2CC"
    CLR_NET_AFTER_POS= "E2EFDA"
    CLR_NET_AFTER_NEG= "FFDCE0"
    CLR_NET_AFTER_ZRO= "FFEB9C"
    CLR_OA_ROW       = "DDEEFF"
    CLR_PLT_ROW      = "EAF4EA"
    CLR_WHITE        = "FFFFFF"
    CLR_GREY_LIGHT   = "F5F5F5"

    COL_WH   = 1
    COL_TYPE = 2
    COL_CP   = 3
    COL_P1   = 4

    def c(col): return get_column_letter(col)

    ws.column_dimensions[c(COL_WH)].width   = 12
    ws.column_dimensions[c(COL_TYPE)].width = 26
    ws.column_dimensions[c(COL_CP)].width   = 10
    for i in range(n_per):
        ws.column_dimensions[c(COL_P1 + i)].width = 8

    # Row 1: capacity header
    row = 1
    cap_val = p.CAP.get(periods[0], 0) if p.CAP else 0
    cap_cell = ws.cell(row=row, column=COL_WH,
                       value=f"Capacity = {int(cap_val)}  |  Product: {p.product}  |  CP={p.CP}")
    cap_cell.font = Font(bold=True, color="FFFFFF")
    cap_cell.fill = _fill(CLR_HEADER_DARK)
    ws.merge_cells(start_row=row, start_column=COL_WH, end_row=row, end_column=COL_CP)

    period_lbl = ws.cell(row=row, column=COL_P1, value="Period")
    period_lbl.font = Font(bold=True, color="FFFFFF")
    period_lbl.fill = _fill(CLR_HEADER_DARK)
    period_lbl.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=COL_P1, end_row=row, end_column=COL_P1 + n_per - 1)

    # Row 2: sub-header
    row = 2
    for col, txt in [(COL_WH, "Warehouse"), (COL_TYPE, "Data Type"), (COL_CP, "Case-pack")]:
        cell = ws.cell(row=row, column=col, value=txt)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill(CLR_HEADER_MID)
        cell.alignment = Alignment(horizontal="center")
    for i, t in enumerate(periods):
        cell = ws.cell(row=row, column=COL_P1 + i, value=t)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill(CLR_HEADER_MID)
        cell.alignment = Alignment(horizontal="center")

    pivot = build_pivot_data(problem, sol)

    # Write warehouse blocks
    current_row = 3

    for wh in p.warehouses:
        wh_data    = pivot[wh]
        n_rows_wh  = len(ROW_LABELS)
        wh_start   = current_row
        wh_end     = current_row + n_rows_wh - 1

        for ri, label in enumerate(ROW_LABELS):
            r = current_row + ri

            wh_cell = ws.cell(row=r, column=COL_WH)
            if ri == 0:
                wh_cell.value = wh
                wh_cell.font  = Font(bold=True)
            wh_cell.fill = _fill(CLR_WH_LABEL)

            type_cell = ws.cell(row=r, column=COL_TYPE, value=label)
            type_cell.font = Font(italic=(label not in ("Net inventory (before)", "Net inventory (after)")))

            cp_cell = ws.cell(row=r, column=COL_CP, value=p.CP)
            cp_cell.alignment = Alignment(horizontal="center")

            for ci, t in enumerate(periods):
                val  = wh_data[label].get(t, 0.0)
                cell = ws.cell(row=r, column=COL_P1 + ci, value=val)
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.##"

                if label == "Net inventory (before)":
                    cell.fill = _fill(CLR_NET_BEFORE)
                    type_cell.fill = _fill(CLR_NET_BEFORE)
                elif label == "Net inventory (after)":
                    if val > 0:
                        cell.fill = _fill(CLR_NET_AFTER_POS)
                    elif val < 0:
                        cell.fill = _fill(CLR_NET_AFTER_NEG)
                    else:
                        cell.fill = _fill(CLR_NET_AFTER_ZRO)
                    type_cell.fill = _fill(CLR_NET_AFTER_POS if val > 0 else (CLR_NET_AFTER_NEG if val < 0 else CLR_NET_AFTER_ZRO))
                    cell.font = Font(bold=True)
                elif "OA" in label:
                    cell.fill = _fill(CLR_OA_ROW)
                elif "PLT" in label:
                    cell.fill = _fill(CLR_PLT_ROW)
                else:
                    cell.fill = _fill(CLR_WHITE if ri % 2 == 0 else CLR_GREY_LIGHT)

        if n_rows_wh > 1:
            ws.merge_cells(start_row=wh_start, start_column=COL_WH, end_row=wh_end, end_column=COL_WH)
            wh_cell_merged = ws.cell(row=wh_start, column=COL_WH)
            wh_cell_merged.alignment = Alignment(vertical="center", horizontal="center", wrap_text=False)

        thick = Side(style="medium")
        thin  = Side(style="thin")
        for ri2 in range(n_rows_wh):
            r2 = wh_start + ri2
            for col2 in range(COL_WH, COL_P1 + n_per):
                cell2 = ws.cell(row=r2, column=col2)
                top_s    = thick if ri2 == 0            else thin
                bottom_s = thick if ri2 == n_rows_wh-1 else thin
                left_s   = thick if col2 == COL_WH     else thin
                right_s  = thick if col2 == COL_P1+n_per-1 else thin
                cell2.border = Border(left=left_s, right=right_s, top=top_s, bottom=bottom_s)

        current_row = wh_end + 1

    current_row += 1
    labels_costs = [
        ("Grand Total Cost",    summary["grand_total_cost"]),
        ("  Overstock",         summary["total_overstock_cost"]),
        ("  Shortage",          summary["total_shortage_cost"]),
        ("  Backorder",         summary["total_backorder_cost"]),
        ("  Penalty OA",        summary["total_penalty_OA_cost"]),
        ("  Transport",         summary["total_transport_cost"]),
        ("Elapsed (s)",         summary["elapsed_s"]),
    ]
    for lbl, val in labels_costs:
        ws.cell(row=current_row, column=COL_WH, value="").fill  = _fill("F0F0F0")
        ws.cell(row=current_row, column=COL_TYPE, value=lbl).font = Font(bold=lbl.startswith("G"))
        ws.cell(row=current_row, column=COL_P1, value=val).number_format = "#,##0.00"
        current_row += 1

    ws.freeze_panes = ws.cell(row=3, column=COL_P1)


def save_excel(wb: openpyxl.Workbook, path: Path) -> None:
    if not HAS_OPENPYXL:
        return
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(str(path))
    log.info("Saved Excel: %s", path)


# ---------------------------------------------------------------------------
# Main Runner for Heuristic Pipeline
# ---------------------------------------------------------------------------

def main() -> None:
    global DATA_DIR
    parser = argparse.ArgumentParser(description="Run Greedy Operational Heuristic for products in inventory_flow.csv")
    parser.add_argument("--data-dir", type=str, default=str(DATA_DIR), help="Path to data directory")
    parser.add_argument("--max-products", type=int, default=999999, help="Max products to run (default all)")
    parser.add_argument("--no-excel", action="store_true", help="Skip Excel output")
    args = parser.parse_args()

    DATA_DIR = Path(args.data_dir)
    if not DATA_DIR.exists():
        log.error("Data directory not found: %s", DATA_DIR)
        sys.exit(1)

    log.info("Loading CSV data from %s...", DATA_DIR.name)
    product_ids, flow_data = load_inventory_flow(n=args.max_products)
    bi_data     = load_inventory_begin()
    costs_data  = load_unit_costs()
    cap_data    = load_vendor_capacity()
    cp_data     = load_packing_details()
    lt_plt      = load_plt_lead_times()
    dist_data   = load_distances()
    LT_OA       = 8
    TC          = 1.2

    log.info("Found %d products to process.", len(product_ids))

    # Prepare output baseline directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    schedule_path = get_writable_path(OUTPUT_DIR / "schedule.csv")
    costs_path    = get_writable_path(OUTPUT_DIR / "costs.csv")
    summary_path  = get_writable_path(OUTPUT_DIR / "cost_summary.csv")
    log_path      = get_writable_path(OUTPUT_DIR / "run_log.csv")
    excel_path    = get_writable_path(OUTPUT_DIR / "schedule_detail.xlsx")

    # Clear existing files (if writable)
    for p in [schedule_path, costs_path, summary_path, log_path, excel_path]:
        try:
            if p.exists():
                p.unlink()
        except Exception as e:
            log.warning("Could not clear existing file %s: %s", p.name, e)

    write_excel = HAS_OPENPYXL and not args.no_excel
    if write_excel:
        wb = openpyxl.Workbook()
    else:
        wb = None

    all_summaries: List[Dict] = []
    success = 0
    failed  = 0

    for idx, pid in enumerate(product_ids, 1):
        log.info("=" * 60)
        log.info("[%d/%d] Heuristic Product: %s", idx, len(product_ids), pid)
        log.info("=" * 60)

        try:
            pflow   = flow_data.get(pid, {})
            bi_map  = bi_data.get(pid, {})
            c_map   = costs_data.get(pid, {})
            cap_map = cap_data.get(pid, {})
            cp      = cp_data.get(pid, 1)

            if not pflow:
                log.warning("  No inventory_flow data for %s — skipping.", pid)
                failed += 1
                continue

            if not bi_map:
                log.warning("  No inventory_begin data for %s — skipping.", pid)
                failed += 1
                continue

            problem = build_problem(
                product_id = pid,
                flow       = pflow,
                bi_map     = bi_map,
                costs_map  = c_map,
                cap_map    = cap_map,
                cp         = cp,
                lt_oa      = LT_OA,
                lt_plt_map = lt_plt,
                dist_map   = dist_data,
                tc         = TC,
            )

            log.info("  %d warehouses | %d periods | CP=%d", problem.n_wh, problem.n_periods, problem.CP)

            t0 = time.perf_counter()
            solver = GreedyOperationalHeuristic(problem)
            sol_dict = solver.solve()
            elapsed = time.perf_counter() - t0

            # Wrap dict solution to behave exactly like the GA solution
            sol_wrapped = HeuristicSolutionWrapper(problem, sol_dict)

            log.info("  Done: cost=%.4f | elapsed=%.4fs", sol_wrapped.fitness, elapsed)

            sched_rows, cost_rows, summary = extract_results(problem, sol_wrapped, elapsed)

            # Append to flat CSVs
            write_csv(schedule_path, sched_rows, mode="a")
            write_csv(costs_path,    cost_rows,  mode="a")
            all_summaries.append(summary)

            # Add to Excel workbook
            if write_excel:
                add_product_sheet(wb, problem, sol_wrapped, summary)
                if idx % 10 == 0 or idx == len(product_ids):
                    save_excel(wb, excel_path)

            log.info("  Overstock : %12.2f", summary["total_overstock_cost"])
            log.info("  Shortage  : %12.2f", summary["total_shortage_cost"])
            log.info("  Backorder : %12.2f", summary["total_backorder_cost"])
            log.info("  Penalty OA: %12.2f", summary["total_penalty_OA_cost"])
            log.info("  Transport : %12.2f", summary["total_transport_cost"])
            log.info("  TOTAL     : %12.2f", summary["grand_total_cost"])

            success += 1

        except Exception as e:
            log.error("  FAILED for %s: %s", pid, e)
            log.debug(traceback.format_exc())
            failed += 1
            all_summaries.append({
                "product"              : pid,
                "n_warehouses"         : 0,
                "n_periods"            : 0,
                "fitness"              : None,
                "elapsed_s"            : None,
                "total_overstock_cost" : None,
                "total_shortage_cost"  : None,
                "total_backorder_cost" : None,
                "total_penalty_OA_cost": None,
                "total_transport_cost" : None,
                "grand_total_cost"     : None,
            })

    if all_summaries:
        write_csv(summary_path, all_summaries, mode="w")
        log.info("Saved cost summary: %s", summary_path)

    log_rows = []
    for s in all_summaries:
        log_rows.append({
            "product"        : s["product"],
            "n_warehouses"   : s["n_warehouses"],
            "n_periods"      : s["n_periods"],
            "fitness"        : s["fitness"],
            "elapsed_s"      : s["elapsed_s"],
            "grand_total_cost": s["grand_total_cost"],
            "status"         : "ok" if s["fitness"] is not None else "failed",
        })
    write_csv(log_path, log_rows, mode="w")

    if write_excel:
        save_excel(wb, excel_path)

    log.info("=" * 60)
    log.info("DONE HEURISTIC PIPELINE: %d success, %d failed", success, failed)
    log.info("Baseline outputs in: %s", OUTPUT_DIR)
    log.info("  schedule.csv         — full schedule (flat CSV)")
    log.info("  costs.csv            — per-row cost breakdown")
    log.info("  cost_summary.csv     — total + component costs per product")
    log.info("  run_log.csv          — timing and fitness summary")
    if write_excel:
        log.info("  schedule_detail.xlsx — pivot schedule, 1 sheet/product")


if __name__ == "__main__":
    main()